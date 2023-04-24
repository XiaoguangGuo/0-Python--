import glob
import os
import tkinter as tk
from tkinter import filedialog
import tkinter.dnd as dnd
from tkinterdnd2 import DND_FILES, TkinterDnD
class DndHandler:
    def __init__(self, widget, on_drop=None):
        self.widget = widget
        self.on_drop = on_drop



def add_image():
    file_path = filedialog.askopenfilename()
    # 在此处处理图像文件并命名


def add_template():
    file_path = filedialog.askopenfilename(
        filetypes=[('Excel files', '*.xlsx')])
    # 在此处处理模板文件并命名


def generate_invoice():
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    import datetime
    import os
    import shutil

    filename = r'D:\\运营\invoice\shipment.xlsx'
    wb = load_workbook(filename)
    sheet = wb.active

    # 初始化变量和结果dataframe
    shipment_data = {
        "Shipment ID": [],
        "Box ID": [],
        "SKU": [],
        "ASIN": [],
        "FNSKU": [],
        "数量": [],
        "Weight": [],
        "Unit": [],
        "Length": [],
        "Width": [],
        "Height": [],
        "UnitDimensions": []
    }
    shipment_id = ""
    box_id = ""
    sku = ""
    asin = ""
    fnsku = ""
    address = ""
    shipment_name = ""

    # 遍历第一列
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            shipment_id = row[0]
            box_id = ""
            sku = ""
            asin = ""
            fnsku = ""
            address = ""
            shipment_name = ""
        else:
            box_id = row[1]
            sku = row[2]
            asin = row[3]
            fnsku = row[4]
            address = row[5]
            shipment_name = row[6]

        # 处理SKU
        if sku is not None:
            sku = sku.strip()
            if sku.startswith("X"):
                sku = sku[1:]

        # 处理ASIN
        if asin is not None:
            asin = asin.strip()
            if asin.startswith("X"):
                asin = asin[1:]

        # 处理FNSKU
        if fnsku is not None:
            fnsku = fnsku.strip()
            if fnsku.startswith("X"):
                fnsku = fnsku[1:]

        # 处理地址
        if address is not None:
            address = address.strip()
            if address.startswith("X"):
                address = address[1:]

        # 处理Shipment ID
        if shipment_id is not None:
            shipment_id = shipment_id.strip()
            if shipment_id.startswith("X"):
                shipment_id = shipment_id[1:]

        # 处理Box ID
        if box_id is not None:
            box_id = box_id.strip()
            if box_id.startswith("X"):
                box_id = box_id[1:]

        # 处理Shipment Name
        if shipment_name is not None:
            shipment_name = shipment_name.strip()
            if shipment_name.startswith("X"):
                shipment_name = shipment_name[1:]

        # 处理重量
        weight = 0
        if sku is not None:
            weight = get_weight(sku)

        # 处理尺寸
        length = 0
        width = 0
        height = 0
        if sku is not None:
            length, width, height = get_dimensions(sku)

        # 处理单位尺寸
        unit_dimensions = ""
        if length > 0 and width > 0 and height > 0:
            unit_dimensions = f"{length}x{width}x{height} in"

        # 添加数据到结果dataframe
        shipment_data["Shipment ID"].append(shipment_id)
        shipment_data["Box ID"].append(box_id)
        shipment_data["SKU"].append(sku)
        shipment_data["ASIN"].append(asin)
        shipment_data["FNSKU"].append(fnsku)
        shipment_data["数量"].append(1)
        shipment_data["Weight"].append(weight)
        shipment_data["Unit"].append("lb")
        shipment_data["Length"].append(length)
        shipment_data["Width"].append(width)
        shipment_data["Height"].append(height)
        shipment_data["UnitDimensions"].append(unit_dimensions)

    # 将结果dataframe写入Excel文件
    df = pd.DataFrame(shipment_data)
    df.to_excel("output.xlsx", index=False)


def view_invoice_history():
    invoice_history_folder = "D:/运营/HistorialData/历史发票"
    invoice_files = glob.glob(os.path.join(invoice_history_folder, "*"))

    # 按时间倒序列出文件
    invoice_files.sort(key=os.path.getmtime, reverse=True)

    # 创建一个新窗口来显示历史发票列表
    history_window = tk.Toplevel(root)
    history_window.title("历史发票")

    def open_invoice(file_path):
        # 用默认程序打开文件
        os.startfile(file_path)

    # 遍历历史发票文件并添加到列表中
    for invoice_file in invoice_files:
        file_name = os.path.basename(invoice_file)
        file_button = tk.Button(
            history_window, text=file_name, command=lambda f=invoice_file: open_invoice(f))
        file_button.pack()

# 其他代码...


def on_drop_image(event):
    widget = event.widget
    if widget != drop_image_frame:
        return

    file_path = root.tk.splitlist(event.data)
    # 在此处处理图像文件并命名


def on_drop_file(event):
    widget = event.widget
    if widget != drop_file_frame:
        return

    file_path = root.tk.splitlist(event.data)
    # 在此处处理文件并命名


root = tk.Tk()
root.title("发票制作")

# 创建菜单栏
menu = tk.Menu(root)
root.config(menu=menu)

# 添加菜单项
filemenu = tk.Menu(menu)
menu.add_cascade(label="广告", menu=filemenu)
menu.add_cascade(label="关键词", menu=filemenu)
menu.add_cascade(label="绩效", menu=filemenu)
menu.add_cascade(label="规则", menu=filemenu)
menu.add_cascade(label="产品管理", menu=filemenu)

invoice_menu = tk.Menu(menu)
menu.add_cascade(label="发票制作", menu=invoice_menu)

# 添加发票制作选项
invoice_menu.add_command(label="站点")
invoice_menu.add_command(label="渠道")
invoice_menu.add_command(label="发货日期")
invoice_menu.add_command(label="目的地")

# 空白输入框
input_text = tk.Text(root, width=50, height=25)
input_text.pack()

# 拖拽图片框
drop_image_frame = tk.Frame(root, width=300, height=100, bg="grey")
drop_image_frame.pack()
drop_image_frame.label = tk.Label(
    drop_image_frame, text="拖拽图片到此处", fg="white", bg="grey")
drop_image_frame.label.pack(expand=True, fill="both")
dnd.DndHandler(drop_image_frame, on_drop=on_drop_image)

# 拖拽文件框
drop_file_frame = tk.Frame(root, width=300, height=100, bg="lightblue")
drop_file_frame.pack()
drop_file_frame.label = tk.Label(
    drop_file_frame, text="拖拽文件到此处", fg="black", bg="lightblue")
drop_file_frame.label.pack(expand=True, fill="both")
dnd.DndHandler(drop_file_frame, on_drop=on_drop_file)

# 增加模板按钮
add_template_button = tk.Button(root, text="增加模板", command=add_template)
add_template_button.pack()

# 生成发票按钮
generate_invoice_button = tk.Button(
    root, text="生成发票", command=generate_invoice)
generate_invoice_button.pack()

# 历史发票按钮
view_invoice_history_button = tk.Button(
    root, text="历史发票", command=view_invoice_history)
view_invoice_history_button.pack()

root.mainloop()
