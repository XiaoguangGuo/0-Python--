import openpyxl

def find_blank_area(file_path, extra_rows=20):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    blank_area_start = -1

    consecutive_blank_rows = 0
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + extra_rows - 5):
        for col in range(1, max_col + 1):
            # 判断当前单元格是否为空白单元格且不是合并单元格
            cell = ws.cell(row=row, column=col)
            if not cell.is_merged and cell.value is None:
                # 判断是否连续8列都是空白单元格且不是合并单元格
                if all([ws.cell(row=row+i, column=col+j).value is None and not ws.cell(row=row+i, column=col+j).is_merged for i in range(6) for j in range(8)]):
                    blank_area_start = row
                    break
        if blank_area_start > 0:
            break

    if blank_area_start > 0:
        col_names = [ws.cell(row=blank_area_start - 1, column=col).value for col in range(1, max_col + 1)]
        top_left_cell_coord = (blank_area_start, col_names.index(col_names[0]) + 1)
        print("Top left cell coordinates: ", top_left_cell_coord)
        print(col_names)
        return top_left_cell_coord, col_names
    else:
        return None
