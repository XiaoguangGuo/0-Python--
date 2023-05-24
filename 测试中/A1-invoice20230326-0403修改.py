import pandas as pd
import numpy as np
from openpyxl import load_workbook

filename = r'D:\\运营\\invoice\\shipment.xlsx'
wb = load_workbook(filename)
sheet = wb.active

# 初始化变量和结果dataframe
shipment_data = {
    "Shipment ID": [],
    "Box ID": [],
    "SKU": [],
    "ASIN": [],
    "FNSKU": [],
    "数量": [],
    "Weight": [],
    "Unit": [],
    "Length": [],
    "Width": [],
    "Height": [],
    "UnitDimensions": []
}
shipment_id = ""
box_id = ""
sku = ""
asin = ""
fnsku = ""

# 遍历第一列
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    if cell_value is not None:
        cell_value = str(cell_value)  # 将cell_value转换为字符串

        if "Shipment ID:" in cell_value:
            shipment_id = cell_value.replace("Shipment ID:", "").strip()

        elif "Box ID:" in cell_value:
            box_id = cell_value.replace("Box ID:", "").strip()

            # 寻找SKU、ASIN和FNSKU以及数量
            sku_row = row + 1
            while sku_row <= sheet.max_row:
                sku_value = str(sheet.cell(row=sku_row, column=1).value)
                if "SKU:" in sku_value:
                    sku = sku_value.replace("SKU:", "").strip()
                    
                    # 查找ASIN
                    asin_row = sku_row + 1
                    while asin_row <= sheet.max_row:
                        asin_value = str(sheet.cell(row=asin_row, column=1).value)
                        if "ASIN:" in asin_value:
                            asin = asin_value.replace("ASIN:", "").strip()
                            break
                        asin_row += 1
                    
                    # 查找FNSKU
                    fnsku_row = sku_row + 1
                    while fnsku_row <= sheet.max_row:
                        fnsku_value = str(sheet.cell(row=fnsku_row, column=1).value)
                        if "FNSKU:" in fnsku_value:
                            fnsku = fnsku_value.replace("FNSKU:", "").strip()
                            break
                        fnsku_row += 1
                    
                    # 查找数量
                    quantity_row = sku_row + 1
                    while quantity_row <= sheet.max_row:
                        try:
                            quantity = int(sheet.cell(row=quantity_row, column=1).value)
                            break
                        except (ValueError, TypeError):
                            quantity_row += 1

                        # 如果当前行是最后一行，寻找下一个有效的数字作为数量
                        if quantity_row == sheet.max_row and quantity is None:
                            search_row = sheet.max_row
                            while search_row > 0:
                                try:
                                    quantity = int(sheet.cell(row=search_row, column=1).value)
                                    break
                                except (ValueError, TypeError):
                                    search_row -= 1
                    

                    weight_row = row + 1
                    dimensions_row = row + 1
                    while weight_row < sheet.max_row:
                        weight_value = str(sheet.cell(row=weight_row, column=1).value)
                        if "weight:" in weight_value.lower():
                            weight = sheet.cell(row=weight_row + 1, column=1).value
                            unit = sheet.cell(row=weight_row + 2, column=1).value
                            break
                        weight_row += 1

                    while dimensions_row < sheet.max_row:
                        dimensions_value = str(sheet.cell(row=dimensions_row, column=1).value)
                        if "dimensions:" in dimensions_value.lower():
                            length = sheet.cell(row=dimensions_row + 1, column=1).value
                            width = sheet.cell(row=dimensions_row + 3, column=1).value
                            height = sheet.cell(row=dimensions_row + 5, column=1).value
                            unit_dimensions = sheet.cell(row=dimensions_row + 6, column=1).value
                            break
                        dimensions_row += 1
                    if "US" in Country  or "NEW-US" in Country:
                        weight = round(weight * 0.45359237, 2)
                        length = round(length * 2.54, 2)
                        width = round(width * 2.54, 2)
                        height = round(height * 2.54, 2)
                        
                    shipment_data["Shipment ID"].append(shipment_id)
                    shipment_data["Shipment ID"].append(shipment_id)
                    shipment_data["Box ID"].append(box_id)
                    shipment_data["SKU"].append(sku)
                    shipment_data["ASIN"].append(asin)
                    shipment_data["FNSKU"].append(fnsku)
                    shipment_data["数量"].append(quantity)
                    shipment_data["Weight"].append(weight)
                    shipment_data["Unit"].append(unit)
                    shipment_data["Length"].append(length)
                    shipment_data["Width"].append(width)
                    shipment_data["Height"].append(height)
                    shipment_data["UnitDimensions"].append(unit_dimensions)

                    sku_row = quantity_row  # 更新sku_row以跳过已处理的数量行
                elif "Box ID:" in sku_value:
                    break  # 跳出循环以处理下一个Box ID
                
                else:
                    sku_row += 1

shipment_df = pd.DataFrame(shipment_data)
shipment_df["图片"]=""
shipment_df["产品链接"]="https://www.amazon.com/dp/"+shipment_df["ASIN"]

print(shipment_df)
shipment_df.to_excel(r'D:\\运营\\invoicedraft.xlsx')


# 读取产品信息表格
# 读取产品信息表格
products_df = pd.read_excel(r'D:\\运营\invoice\\发票基础信息表.xlsx')
products_df['SKU'] = products_df['SKU'].fillna('')

# 读取另一个 DataFrame
other_df = shipment_df
 
# 遍历 other_df 的 SKU 列，在 products_df 的产品 SKU 列中找到包含对应值的单元格，将这一行对应的信息合并到 other_df 中
for index, row in other_df.iterrows():
    for sku in row['SKU'].split():
        matching_rows = products_df[products_df['SKU'].str.contains(sku)]
        if not matching_rows.empty:
            matching_row = matching_rows.iloc[0]
            for col_name in ['产品SKU', '英文产品名称', '中文产品名称', '单个产品申报价值USD', "HSCODE",
                             'Brand(品牌)*', 'Model（型号）*', '中文材质',
                             'Purpose(用途)*', '是否带电（Y/N）', 'PICTURES（图片）*',
                             '产品销售链接', '内部名称', '产品销售价格', '英文材质',
                             '英文用途']:
                other_df.at[index, col_name] = matching_row[col_name]
            other_df.at[index, 'SKU'] = matching_row['SKU']
            
 
print(other_df.columns)

# 修改箱尺寸列
other_df["箱尺寸"] = other_df["Length"].astype(str) + "*" + other_df["Width"].astype(str) + "*" + other_df["Height"].astype(str)

# 计算箱体积并保留小数后三位
other_df["箱体积"] = (other_df["Length"] * other_df["Width"] * other_df["Height"] / 1000000).round(3)

# 增加箱数列
other_df["箱数"] = 1

# 修改中英文材质列
other_df["中英文材质"] = other_df["中文材质"] + ' ' + other_df["英文材质"]

# 合并中英文用途列
other_df["中英文用途"] = other_df["Purpose(用途)*"] + other_df["英文用途"]

# 计算总申报价
other_df["总申报价"] = other_df["单个产品申报价值USD"] * other_df["数量"]# 将合并后的 DataFrame 写入 Excel 文件
 



assert "Box ID" in other_df.columns, "Box_ID 列不存在"
assert "箱数" in other_df.columns, "箱数 列不存在"
assert "箱体积" in other_df.columns, "箱体积 列不存在"
assert "箱尺寸" in other_df.columns, "箱尺寸 列不存在"

# 遍历数据框的行
for index, row in other_df.iterrows():
    # 如果当前行不是第一行且 Box_ID 与上一行的 Box_ID 相同
    if index > 0 and other_df.at[index, "Box ID"] == other_df.at[index - 1, "Box ID"]:
        # 将 箱数, 箱体积, 箱尺寸 列设置为空值
        other_df.at[index, "箱数"] = ""
        other_df.at[index, "箱体积"] = ""
        other_df.at[index, "箱尺寸"] = ""

other_df.to_excel(r'D:\\运营\invoice\\invoicedraft2.xlsx')

# 打印合并后的 DataFrame
print(other_df)

import pandas as pd
import openpyxl

import openpyxl

def find_blank_area(file_path, extra_rows=10):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]
    blank_area_start = -1

    consecutive_blank_rows = 0
    max_row = ws.max_row
    max_col = ws.max_column
    print(max_row, max_col)
    for row in range(1, max_row + extra_rows + 1):
        empty_cells = sum([1 for col in range(1, max_col + 1) if ws.cell(row=row, column=col).value is None and ws.cell(row=row, column=col) not in ws.merged_cells])

        if empty_cells >= 6:
            consecutive_blank_rows += 1
        else:
            consecutive_blank_rows = 0

        if consecutive_blank_rows == 6:
            blank_area_start = row - 5
            break

    if blank_area_start > 0:
        col_names = [ws.cell(row=blank_area_start - 1, column=col).value for col in range(1, max_col + 1) if ws.cell(row=blank_area_start - 1, column=col) not in ws.merged_cells]
        print(col_names)
        return col_names
    else:
        return None
    
def match_col_names(template_col_names, mapping_file_path, mapping_sheet_name):
    mapping_df = pd.read_excel(mapping_file_path, sheet_name=mapping_sheet_name)
    middle_col_names = []

    for col_name in template_col_names:
        middle_col_name = mapping_df.loc[mapping_df['模板列名'] == col_name, '中间列名']
        if not middle_col_name.empty:
            middle_col_names.append(middle_col_name.iloc[0])
        else:
            middle_col_names.append(None)

    return middle_col_names

if __name__ == "__main__":
    file_path = "D:\\运营\\发票模板\\template.xlsx"
    mapping_file_path = "D:\\运营\\invoice\\发票基础信息表.xlsx"
    mapping_sheet_name = "列名对应"
    
    template_col_names = find_blank_area(file_path)
    if template_col_names:
        middle_col_names = match_col_names(template_col_names, mapping_file_path, mapping_sheet_name)
        print("转换后的中间列名列表：", middle_col_names)
    else:
        print("在工作簿中未找到满足条件的空白区域。")


new_col_names = []
tbd_counter = 1
for col in middle_col_names:
    if pd.isna(col):
        new_col_names.append(f'tobedecided{tbd_counter}')
        tbd_counter += 1
    else:
        new_col_names.append(col)

# 选择与new_col_names中的列名匹配的列
other_df_selected = other_df[[col for col in new_col_names if col in other_df.columns]]

# 添加new_col_names中不存在于other_df的列，用空值填充
for col in new_col_names:
    col_str = str(col)  # 将列名转换为字符串
    if col not in other_df.columns:
        other_df_selected = other_df_selected.assign(**{col_str: ''})

# 将other_df_selected的列顺序调整为与new_col_names相同
other_df_selected = other_df_selected[new_col_names]
other_df_selected.to_excel(r'D:\\运营\invoice\\invoicedraft-new.xlsx')
