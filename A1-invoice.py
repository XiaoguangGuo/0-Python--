import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
import os
import shutil


# 请输入国家名
Shipment_dic = {}
tempplate_file_path ="D:\\运营\\发票模板\\"
country=input("请输入国家名:US,CA,JP，NEW-US,NEW-CA,UK,IT")
link_dic={"US":"www.amazon.com/dp/","JP":"www.amazon.co.jp/dp/","UK":"www.amazon.co.uk/dp/","CA":"www.amazon.ca/dp/"}
#输入渠道名
channel=input("请输入渠道名:全中,小平")
#遍历D:\运营\发票模板\文件夹下的所有文件，找到第一个包含国家名和渠道名的文件并用load_workbook打开


found_file = False

while not found_file:
    for root, dirs, files in os.walk(tempplate_file_path):
        for file in files:
            if country in file and channel in file:
                print(file)
                templatefile = tempplate_file_path + file
                found_file = True
                break
        if found_file:
            break

    if not found_file:
        print("没有模板，请先设置模板")
        user_input = input("设置好后按回车继续，或输入 'S' 退出程序：")
        if user_input.lower() == "s":
            sys.exit("程序已退出")

             
filename = r'D:\\运营\invoice\shipment.xlsx'
wb = load_workbook(filename)
sheet = wb.active

# 初始化变量和结果dataframe
shipment_data = {
    "Shipment ID": [],
    "Box ID": [],
    "SKU": [],
    "ASIN": [],
    "FNSKU": [],
    "数量": [],
    "Weight": [],
    "Unit": [],
    "Length": [],
    "Width": [],
    "Height": [],
    "UnitDimensions": []
}
shipment_id = ""
box_id = ""
sku = ""
asin = ""
fnsku = ""
address = ""
shipment_name = ""

# 遍历第一列
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    if cell_value is not None:
        cell_value = str(cell_value)  # 将cell_value转换为字符串

        if "Shipment ID:" in cell_value:
            shipment_id = cell_value.replace("Shipment ID:", "").strip()
        elif  "Amazon Reference ID:" in cell_value:
            Reference_id = cell_value.replace("Amazon Reference ID:", "").strip()
        elif "Ship to:" in cell_value:
            address = cell_value.replace("Ship to:", "").strip()
            Shipment_dic["Address"] = address
        elif "Shipment name:" in cell_value:
            shipment_name = cell_value.replace("Shipment name:", "").strip()
            Shipment_dic["Shipment Name"] = shipment_name
        elif "Box ID:" in cell_value:
            box_id = cell_value.replace("Box ID:", "").strip()

            # 寻找SKU、ASIN和FNSKU以及数量
            sku_row = row + 1
            while sku_row <= sheet.max_row:
                sku_value = str(sheet.cell(row=sku_row, column=1).value)
                if "SKU:" in sku_value:
                    sku = sku_value.replace("SKU:", "").strip()
                    
                    # 查找ASIN
                    asin_row = sku_row + 1
                    while asin_row <= sheet.max_row:
                        asin_value = str(sheet.cell(row=asin_row, column=1).value)
                        if "ASIN:" in asin_value:
                            asin = asin_value.replace("ASIN:", "").strip()
                            break
                        asin_row += 1
                    
                    # 查找FNSKU
                    fnsku_row = sku_row + 1
                    while fnsku_row <= sheet.max_row:
                        fnsku_value = str(sheet.cell(row=fnsku_row, column=1).value)
                        if "FNSKU:" in fnsku_value:
                            fnsku = fnsku_value.replace("FNSKU:", "").strip()
                            break
                        fnsku_row += 1
                    
                    # 查找数量
                    quantity_row = sku_row + 1
                    while quantity_row <= sheet.max_row:
                        try:
                            quantity = int(sheet.cell(row=quantity_row, column=1).value)
                            break
                        except (ValueError, TypeError):
                            quantity_row += 1

                        # 如果当前行是最后一行，寻找下一个有效的数字作为数量
                        if quantity_row == sheet.max_row and quantity is None:
                            search_row = sheet.max_row
                            while search_row > 0:
                                try:
                                    quantity = int(sheet.cell(row=search_row, column=1).value)
                                    break
                                except (ValueError, TypeError):
                                    search_row -= 1
                    

                    weight_row = row + 1
                    dimensions_row = row + 1
                    while weight_row < sheet.max_row:
                        weight_value = str(sheet.cell(row=weight_row, column=1).value)
                        if "weight:" in weight_value.lower():
                            weight = sheet.cell(row=weight_row + 1, column=1).value
                            unit = sheet.cell(row=weight_row + 2, column=1).value
                            break
                        weight_row += 1

                    while dimensions_row < sheet.max_row:
                        dimensions_value = str(sheet.cell(row=dimensions_row, column=1).value)
                        if "dimensions:" in dimensions_value.lower():
                            length = sheet.cell(row=dimensions_row + 1, column=1).value
                            width = sheet.cell(row=dimensions_row + 3, column=1).value
                            height = sheet.cell(row=dimensions_row + 5, column=1).value
                            unit_dimensions = sheet.cell(row=dimensions_row + 6, column=1).value
                            break
                        dimensions_row += 1

                    shipment_data["Shipment ID"].append(shipment_id)
                    
                    shipment_data["Box ID"].append(box_id)
                    shipment_data["SKU"].append(sku)
                    shipment_data["ASIN"].append(asin)
                    shipment_data["FNSKU"].append(fnsku)
                    shipment_data["数量"].append(quantity)
                    shipment_data["Weight"].append(weight)
                    shipment_data["Unit"].append(unit)
                    shipment_data["Length"].append(length)
                    shipment_data["Width"].append(width)
                    shipment_data["Height"].append(height)
                    shipment_data["UnitDimensions"].append(unit_dimensions)

                    sku_row = quantity_row  # 更新sku_row以跳过已处理的数量行
                elif "Box ID:" in sku_value:
                    break  # 跳出循环以处理下一个Box ID
                
                else:
                    sku_row += 1
#将shipment_data转换为dataframe             
shipment_data["Address"] = address

shipment_data["Shipment Name"] = shipment_name
shipment_data = pd.DataFrame(shipment_data)
shipment_data["Reference ID"]=Reference_id
shipment_data["图片"]=""
shipment_data["产品链接"]=link_dic[country]+shipment_data["ASIN"]
#定义shipment_df的“发货日期”列等于今天的日期
shipment_data["发货日期"]=datetime.date.today()
#定义shipment_df的“Country”列等于country
shipment_data["Country"]=country
#定义shipment_df的“Channel”列等于channel
shipment_data["Channel"]=channel
#定义shipment_dic
Shipment_dic["Shipment ID"] = shipment_id
Shipment_dic["Reference ID"] = Reference_id
Shipment_dic["Country"] = country
Shipment_dic["Channel"] = channel
print(shipment_data)
#to_excel
output_SUmmary_file = r'D:\\运营\invoice\\发货信息汇总表.xlsx'

if os.path.exists(output_SUmmary_file ):
    # 如果文件存在，读取现有文件并将新数据追加到文件中
    existing_df = pd.read_excel(output_SUmmary_file )
    combined_df = pd.concat([existing_df, shipment_data], ignore_index=True)

    with pd.ExcelWriter(output_SUmmary_file , mode='w', engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False)
    print("数据已附加到发货信息汇总表")
else:
    # 如果文件不存在，创建一个新文件并将数据写入其中
    shipment_df.to_excel(output_SUmmary_file , index=False)
    print("生成新的发货信息汇总表")
 
# 读取产品信息表格
products_df = pd.read_excel(r'D:\\运营\invoice\\发票基础信息表.xlsx')
products_df['SKU'] = products_df['SKU'].fillna('')

# 读取另一个 DataFrame
other_df = shipment_data
 
# 遍历 other_df 的 SKU 列，在 products_df 的产品 SKU 列中找到包含对应值的单元格，将这一行对应的信息合并到 other_df 中
for index, row in other_df.iterrows():
    for sku in row['SKU'].split():
        matching_rows = products_df[products_df['SKU'].str.contains(sku)]
        if not matching_rows.empty:
            matching_row = matching_rows.iloc[0]
            for col_name in ['产品简化名', '英文产品名称', '中文产品名称', '单个产品申报价值USD', "HSCODE",
                             'Brand(品牌)*', 'Model（型号）*', '中文材质',
                             'Purpose(用途)*', '是否带电', 'PICTURES（图片）*',
                             '产品销售链接', '内部名称', '产品销售价格', '英文材质',
                             '英文用途',"带磁","是否含液体","是否危险品"]:

                other_df.at[index, col_name] = matching_row[col_name]
            other_df.at[index, 'SKU'] = matching_row['SKU']
            
 
print(other_df.columns)

# 修改箱尺寸列
other_df["箱尺寸"] = other_df["Length"].astype(str) + "*" + other_df["Width"].astype(str) + "*" + other_df["Height"].astype(str)

# 计算箱体积并保留小数后三位
other_df["箱体积"] = (other_df["Length"] * other_df["Width"] * other_df["Height"] / 1000000).round(3)

# 增加箱数列
other_df["箱数"] = 1

# 修改中英文材质列
other_df["中英文材质"] = other_df["中文材质"] + ' ' + other_df["英文材质"]

# 合并中英文用途列
other_df["中英文用途"] = other_df["Purpose(用途)*"] + other_df["英文用途"]

# 计算总申报价
other_df["申报总价"] = other_df["单个产品申报价值USD"] * other_df["数量"]# 将合并后的 DataFrame 写入 Excel 文件
 
other_df["Net Weight"]=other_df["Weight"]*0.9


assert "Box ID" in other_df.columns, "Box_ID 列不存在"
assert "箱数" in other_df.columns, "箱数 列不存在"
assert "箱体积" in other_df.columns, "箱体积 列不存在"
assert "箱尺寸" in other_df.columns, "箱尺寸 列不存在"

# 遍历数据框的行
for index, row in other_df.iterrows():
    # 如果当前行不是第一行且 Box_ID 与上一行的 Box_ID 相同
    if index > 0 and other_df.at[index, "Box ID"] == other_df.at[index - 1, "Box ID"]:
        # 将 箱数, 箱体积, 箱尺寸 列设置为空值
        other_df.at[index, "箱数"] = ""
        other_df.at[index, "箱体积"] = ""
        other_df.at[index, "箱尺寸"] = ""

other_df["Country"]=country
invoiceALL_df=pd.read_excel(r'D:\运营\Invoice\\invoiceALL.xlsx')
invoiceALL_df = pd.concat([invoiceALL_df, other_df], axis=0, ignore_index=True)
invoiceALL_df.to_excel(r'D:\运营\Invoice\\invoiceALL.xlsx',index=False)


# 打印合并后的 DataFrame
print(other_df)
from openpyxl.worksheet.cell_range import CellRange
from openpyxl import load_workbook, utils
def is_merged_cell(ws, row, col):
    cell_range = CellRange(f"{utils.get_column_letter(col)}{row}:{utils.get_column_letter(col)}{row}")
    for merged_cell in ws.merged_cells.ranges:
        if merged_cell.issuperset(cell_range):
            return True
    return False


def find_blank_area(file_path, extra_rows=30):
    wb = load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]
    blank_area_start = -1
    consecutive_blank_rows = 0
    max_row = ws.max_row
    max_col = ws.max_column
    print(max_row, max_col)

    for row in range(1, max_row + extra_rows + 1):
        empty_cells = [col for col in range(1, max_col + 1) if ws.cell(row=row, column=col).value is None and not is_merged_cell(ws, row, col)]

        if len(empty_cells) >= 8:
            consecutive_blank_rows += 1
        else:
            consecutive_blank_rows = 0

        if consecutive_blank_rows == 6:
            blank_area_start = row - 5
            break

    if blank_area_start > 0:
        top_left_cell_coord = (blank_area_start, empty_cells[0])
        print("Top left cell coordinates: ", top_left_cell_coord)

        col_names_row = blank_area_start - 1
        col_names = [ws.cell(row=col_names_row, column=col).value for col in range(1, max_col + 1)]
        print("Column names: ", col_names)
        return top_left_cell_coord, col_names
    else:
        return None, None




def match_col_names(template_col_names, mapping_file_path, mapping_sheet_name):
    mapping_df = pd.read_excel(mapping_file_path, sheet_name=mapping_sheet_name)
    middle_col_names = []

    for col_name in template_col_names:
        middle_col_name = mapping_df.loc[mapping_df['模板列名'] == col_name, '中间列名']
        if not middle_col_name.empty:
            middle_col_names.append(middle_col_name.iloc[0])
        else:
            middle_col_names.append(None)

    return middle_col_names

if __name__ == "__main__":
    file_path =templatefile
    mapping_file_path = "D:\\运营\\invoice\\发票基础信息表.xlsx"
    mapping_sheet_name = "列名对应"
    
    top_left_cell_coord,template_col_names = find_blank_area(file_path)
    if template_col_names:
        middle_col_names = match_col_names(template_col_names, mapping_file_path, mapping_sheet_name)
        print("转换后的中间列名列表：", middle_col_names)
    else:
        print("在工作簿中未找到满足条件的空白区域。")


new_col_names = []
tbd_counter = 1
for col in middle_col_names:
    if pd.isna(col):
        new_col_names.append(f'tobedecided{tbd_counter}')
        tbd_counter += 1
    else:
        new_col_names.append(col)
print(new_col_names)
# 选择与new_col_names中的列名匹配的列
other_df_selected = other_df[[col for col in new_col_names if col in other_df.columns]]


# 添加new_col_names中不存在于other_df的列，用空值填充
for col in new_col_names:
    col_str = str(col)  # 将列名转换为字符串
    if col not in other_df.columns:
        other_df_selected = other_df_selected.assign(**{col_str: ''})
  
# 将other_df_selected的列顺序调整为与new_col_names相同
other_df_selected = other_df_selected[new_col_names]
print(other_df_selected.columns)
other_df_selected.to_excel(r'D:\\运营\invoice\\invoicedraft-new'+country+".xlsx")





today = datetime.datetime.today().strftime('%Y%m%d')
new_file_name = country + today + '.xlsx'
new_file_path = os.path.join(os.path.dirname(tempplate_file_path), new_file_name)
shutil.copyfile(templatefile, new_file_path)

template_workbook = load_workbook(new_file_path )
template_worksheet = template_workbook.active

print(top_left_cell_coord)

# 获取 other_df 的行数和列数
num_rows, num_cols = other_df_selected.shape

# 将 other_df 的数据逐个写入模板文件中
for row in range(num_rows):
    for col in range(num_cols):
        cell_value = other_df_selected.iloc[row, col]
        template_worksheet.cell(row=top_left_cell_coord[0] + row, column=int(top_left_cell_coord[1]) + col, value=cell_value)

# 保存模板文件
template_workbook.save(tempplate_file_path + new_file_name)


