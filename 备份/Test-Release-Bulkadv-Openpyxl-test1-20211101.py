# 使用Oenyxl编写，没有使用pandas
### 第一部分汇总各国bulk广告数据到汇总表
#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。
#此程序判断近期10周的数据，合并到一起。这样数据集不至于过大。
# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import shutil

print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中。此操作将更新周广告汇总表为最新10周以内")
a=input('按任意键继续，如未准备好则取消程序执行', )

newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)

#创建汇总表
#bulkhzWorkbook=Workbook()
#sheet=bulkhzWorkbook.active
#Listcol=[ 'Record ID', 'Record Type','Campaign ID','Campaign','Campaign Daily Budget','Portfolio ID','Campaign Start Date','Campaign End Date','Campaign Targeting Type','Ad Group','Max Bid', 'Keyword or Product Targeting', 'Product Targeting ID', 'Match Type','SKU','Campaign Status','Ad Group Status','Status','Impressions','Clicks',
#'Spend','Orders','Total Units', 'Sales' ,'ACoS','Bidding strategy','Placement Type','Increase bids by placement','Country','日期','周数']
#sheet.append(Listcol)
#bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表NEW.xlsx')


bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
bulkoperationworkbook=load_workbook(r'D:\运营\bulkoperation模板.xlsx')
sheetbulkoperation=bulkoperationworkbook.worksheets[0]
                             

# 从第sheet.max_row+1行将表格New写入。
# 在表格的第  列写入国家名 第  列 写入日期  第  列写入 周数 。

#newdata=sheet

# 用pandas
#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'
for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    

    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
#Newdata=pd_read_excel(r'D:\运营\周bulk广告数据\D:/运营/周Bulk广告数据/GV-CA_bulk-a2ijxz1fd77inr-20210411-20210417-1620645025514.xlsx',row=1,
    #newdata=pd.read_excel(newfilepath+'GV-CA_bulk-a2ijxz1fd77inr-20210411-20210417-1620645025514.xlsx',engine="openpyxl",sheet_name="Sponsored Products Campaigns").assign(Country=os.path.basename(newdataname).split('_')[0], 日期=os.path.basename(newdataname).split('-')[4]) 
#增加周数并赋值1
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=1
            #print(sourcedata)
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
        sheet.append(row)
        sheetbulkoperation.append(row)
                #print(sheet.max_row,sheet.max_column)
        
    
    
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')
    bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。
    #sheet.cell(row=2, column=2).value = "omg"
    #print(sheet.cell(2,2).value)
    shutil.move(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\运营\HistoricalData\周bulk广告数据')

for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=30).value
    #print(type(b),b)
    #sheet.cell(row=i,column=30)=datetime.datetime.fromisoformat('2020-12-12 12:22:22') 这个值得试试。
    #sheet.cell(row=i,column=30)==datetime.datetime.strptime(b,'%Y%m%d')

    #sheet.cell(row=i,column=30).value=datetime.datetime.strptime(b,'%Y%m%d')
    #sheet.cell(row=i,column=30).number_format = 'yyyy-mm-dd hh:mm:ss'
    bnewtime=sheet.cell(row=i,column=30).value
    #print(type(bnewtime),bnewtime)
    #bt=datetime.datetime.strptime(b,'%Y-%m-%d %H:%M:%S')
    #print(bt)
    c=(maxtime-bnewtime).days//7+1 #c=(a-b).days()错误
    #print(maxtime,b,c)     
        #caozuoriqi= datetime.datetime.strptime(sheet.cell(row=i, column=30).value, "%Y%m/%d/")
        #print(caozuoriqi)
    #zhoushu=(('2021-10-30').dt.days-(cell(row=i,column=30).value).dt.days)//7+1
    sheet.cell(row=i, column=31).value =c
for j in range(2,sheetbulkoperation.max_row+1):
    b=sheet.cell(row=j,column=30).value
    bnewtime=sheet.cell(row=i,column=30).value
    c=(maxtime-bnewtime).days//7+1
    sheetbulkoperation.cell(row=i, column=31).value =c                         
    
    
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')
bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')


#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序
                             
print("以下为处理bulk操作报表的程序")

#定义bulk数据汇总表所在路径
Allbulkpath='D:\\运营\\'
Allbulk=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx')

#####Campaign汇总表


AllbulkCampaign=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign"],as_index=False)['Impressions','Clicks','Spend','Orders','Total Units','Sales'].agg('sum')

AllbulkCampaign.to_excel(Allbulkpath+'周bulk数据-Campaign汇总.xlsx',index=False)

##### Campaign周汇总表
AllbulkCampaignWEEK=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign","周数"],as_index=False)['Impressions','Clicks','Spend','Orders','Total Units','Sales'].agg('sum')
AllbulkCampaignWEEK.to_excel(Allbulkpath+'周bulk数据-Campaign-WEEK-汇总.xlsx',index=False)



print(Allbulk)
# AllbulkCampaignKeyword=Allbulk[Allbulk['Keyword or Product Targeting'].notna()].groupby(["Country","Campaign",'Keyword or Product Targeting'],as_index=False)['Impressions','Clicks','Spend','Orders','Total Units','Sales'].agg('sum')
AllbulkCampaignKeyword=Allbulk.groupby(["Country","Campaign",'Keyword or Product Targeting'],as_index=False)['Impressions','Clicks','Spend','Orders','Total Units','Sales'].agg('sum')
print(AllbulkCampaignKeyword)
maxrow=len(AllbulkCampaignKeyword)
AllbulkCampaignKeyword["zhuanhualv"]=AllbulkCampaignKeyword['Orders']/AllbulkCampaignKeyword['Clicks']
print(maxrow)
print(AllbulkCampaignKeyword)
clickstest=AllbulkCampaignKeyword.iloc[[0],[5]].values[0][0]
print(clickstest)
AllbulkCampaignKeyword.to_excel(Allbulkpath+'周bulk广告数据汇总表huizong.xlsx')

bulkoperation=pd.read_excel(Allbulkpath+'bulkoperation.xlsx')
bulkoperation["是否更新"]=""
bulkoperationmaxrow=len(bulkoperation)

for i in range(0,maxrow):
    clicks=AllbulkCampaignKeyword.iloc[[i],[4]].values[0][0]
    zhuanhualv=AllbulkCampaignKeyword.iloc[[i],[9]].values[0][0]
    keyword=AllbulkCampaignKeyword.iloc[[i],[2]].values[0][0]
    #print(keyword)

    #print("clicks",clicks,"zhuanhualv",zhuanhualv)
    
    condition01=clicks>16
    condition02=clicks>9 and clicks<17
    condition11=zhuanhualv<0.0625 and zhuanhualv>0
    condition12=zhuanhualv==0
    condition61=zhuanhualv>0.5
    condition63=zhuanhualv>0.2
    print(condition63)
    

    if condition01==True and condition12==True:
        #修改操作bulk表
        for boj in range(0,bulkoperationmaxrow):
            samekeywordCondition= bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[11]].values[0][0]==keyword
            if samekeywordCondition==True:
                bulkoperation.iloc[boj,17]='paused'
                bulkoperation.iloc[boj,31]='此行更新'
                print("newstatus")
    elif condition02==True and condition11==True:
        for boj in range(0,bulkoperationmaxrow):
            samekeywordCondition= bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[11]].values[0][0]==keyword
            if samekeywordCondition==True:
                bulkoperation.iloc[boj,10]=bulkoperation.iloc[boj,10]*0.33
                bulkoperation.iloc[boj,31]='此行更新'
                print("newprice2")
    elif condition01== True and condition11==True:
        for boj in range(0,bulkoperationmaxrow):
            samekeywordCondition= bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[11]].values[0][0]==keyword
            if samekeywordCondition==True:
                bulkoperation.iloc[boj,10]=bulkoperation.iloc[boj,10]*0.33
                bulkoperation.iloc[boj,31]='此行更新'
                print("newprice1",boj)
    elif condition02==True and condition12==True:
         for boj in range(0,bulkoperationmaxrow):
            samekeywordCondition= bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[11]].values[0][0]==keyword
            if samekeywordCondition==True:
                bulkoperation.iloc[boj,17]='paused'
                bulkoperation.iloc[boj,31]='此行更新'
                print("newstatus",boj)
    elif condition63==True:
        for boj in range(0,bulkoperationmaxrow):
            samekeywordCondition= bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[11]].values[0][0]==keyword
            samecompaignAdCondition=bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and (bulkoperation.iloc[[boj],[1]].values[0][0]=="Campaign" or bulkoperation.iloc[[boj],[1]].values[0][0]=="Ad" or bulkoperation.iloc[[boj],[1]].values[0][0]=="Ad Group")
            samecompaignAdgroupCondition=bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and (bulkoperation.iloc[[boj],[1]].values[0][0]=="Campaign" or bulkoperation.iloc[[boj],[1]].values[0][0]=="Ad Group")
            samecompaignCondition=bulkoperation.iloc[[boj],[28]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[0]].values[0][0] and bulkoperation.iloc[[boj],[3]].values[0][0]==AllbulkCampaignKeyword.iloc[[i],[1]].values[0][0] and bulkoperation.iloc[[boj],[1]].values[0][0]=="Campaign" 
            if samekeywordCondition== True:
                print("samekeywordCondition== True")
                bulkoperation.iloc[boj,17]='enabled'
                bulkoperation.iloc[boj,16]='enabled'
                bulkoperation.iloc[boj,15]='enabled'
                bulkoperation.iloc[boj,31]='此行更新'
            elif samecompaignAdCondition== True:
                print("samecompaignAdCondition== True")
                bulkoperation.iloc[boj,17]='enabled'
                bulkoperation.iloc[boj,16]='enabled'
                bulkoperation.iloc[boj,31]='此行更新'
          
            elif samecompaignAdgroupCondition==True:
                print("samecompaignAdgroupCondition==True")
                bulkoperation.iloc[boj,16]='enabled'
                bulkoperation.iloc[boj,15]='enabled'
                bulkoperation.iloc[boj,31]='此行更新'
            elif samecompaignCondition==True:
                print("samecompaign==True")
                bulkoperation.iloc[boj,15]='enabled'
                bulkoperation.iloc[boj,31]='此行更新'
               
                print("good ad confirmed",boj)
               
            
           #else:
               # 写一行此关键词的广告
            
           
           
        

#将bulkoperation 导出

bulkoperation.to_excel(Allbulkpath+'bulkoperationnew.xlsx',index=False)









      
#for i in range(0,maxrow+1):
   # t=AllbulkCampaignKeyword.iloc[[i],[5]].values[0][0]/AllbulkCampaignKeyword.iloc[[i],[4]].values[0][0]
    #print(t)

#Allbulk_groupbycampaign-keyword=
#bulknow=
#wordi=pd(a,i)
#zhuanhualvi=pd(a,4)/pd(a,5)
#for i=(0,maxrow)
    #if clicksi>7 and zhuanhualvi<0.05
        #for bulknow 遍历列行和列
            #if bulknow(a,i)=wordi and compaign 相等 国家 相等
                #bulknowprice=
                #bulknowprice=1/3
    #elif zhuanhualv>0.2

        #for  bulknow 遍历列行和列
            #if
            #bulknowprice=1.3*bulknowprice
            
       



