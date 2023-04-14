import pandas as pd
import openpyxl
from concurrent.futures import ProcessPoolExecutor


def process_sku_rows(sheet, row, shipment_id, box_id):
    sku_data = {}
    sku_data["Shipment ID"] = shipment_id
    sku_data["Box ID"] = box_id
    sku_data["SKU"] = sheet.cell(row=row, column=1).value.replace("SKU:", "").strip()
    row += 1
    while row <= sheet.max_row and "SKU:" not in str(sheet.cell(row=row, column=1).value):
        cell_value = sheet.cell(row=row, column=1).value
        if "weight:" in str(cell_value).lower():
            sku_data["Weight"] = sheet.cell(row=row + 1, column=1).value
            sku_data["Unit"] = sheet.cell(row=row + 2, column=1).value
        elif "dimensions:" in str(cell_value).lower():
            sku_data["Length"] = sheet.cell(row=row + 1, column=1).value
            sku_data["Width"] = sheet.cell(row=row + 3, column=1).value
            sku_data["Height"] = sheet.cell(row=row + 5, column=1).value
            sku_data["UnitDimensions"] = sheet.cell(row=row + 6, column=1).value
        elif "quantity:" in str(cell_value).lower():
            sku_data["数量"] = sheet.cell(row=row + 1, column=1).value
        row += 1
    return sku_data


def process_box_rows(sheet, row):
    box_data = {}
    box_data["Box ID"] = sheet.cell(row=row, column=1).value.replace("Box ID:", "").strip()
    row += 1
    while row <= sheet.max_row and "SKU:" not in str(sheet.cell(row=row, column=1).value):
        row += 1
    return row, box_data


def process_sheet(sheet):
    shipment_data = []
    shipment_id = sheet.cell(row=1, column=1).value.replace("Shipment ID:", "").strip()
    row = 2
    while row <= sheet.max_row:
        cell_value = sheet.cell(row=row, column=1).value
        if "Box ID:" in str(cell_value):
            row, box_data = process_box_rows(sheet, row)
            box_data["Shipment ID"] = shipment_id
            for sku_row in range(row, sheet.max_row + 1):
                cell_value = sheet.cell(row=sku_row, column=1).value
                if "SKU:" in str(cell_value):
                    sku_data = process_sku_rows(sheet, sku_row, shipment_id, box_data["Box ID"])
                    shipment_data.append(sku_data)
        row += 1
    return shipment_data


def main():
    filename = 'D:\\运营\\1数据源\\invoice\\shipment.xlsx'
    output_filename = 'D:\\运营\\invoicedraft.xlsx'
    wb = openpyxl.load_workbook(filename)

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            shipment_data = process_sheet(sheet)
            shipment_df = pd.DataFrame(shipment_data)
            shipment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("程序执行完毕")


if __name__ == '__main__':
    main()
