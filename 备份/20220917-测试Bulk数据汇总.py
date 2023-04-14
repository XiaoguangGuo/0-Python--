# 使用Oenyxl编写，没有使用pandas
### 第一部分汇总各国bulk广告数据到汇总表
#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。

# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表


print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )

newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)


#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#？？？旧的程序，不要了bulkoperationworkbook=load_workbook(r'D:\运营\bulkoperation模板.xlsx')
#???sheetbulkoperation=bulkoperationworkbook.worksheets[0]

#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'

for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta) 
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
    sourcedata.replace(",",".",inplace=True)
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=1
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx') #汇总所有广告数据
#???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

#拷贝一份sourcedata到bulkoperation文件夹,文件名不变
    shutil.copy(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\\运营\\bulkoperationfiles\\')
#移动广告数据到历史数据 
    shutil.move(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile),r'D:\\运营\\HistoricalData\\周bulk广告数据\\')
    
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。


for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=30).value

    bnewtime=sheet.cell(row=i,column=30).value

    c=(maxtime-bnewtime).days//7+1

    sheet.cell(row=i, column=31).value =c
    
                      
    
#保存所有广告数据汇总表
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')
###???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

