import pandas as pd
from openpyxl import load_workbook


def process_sheet(sheet, sku_row, shipment_id, box_id, shipment_data):
    sku_value = str(sheet.cell(row=sku_row, column=1).value)
    if "SKU:" in sku_value:
        sku = sku_value.replace("SKU:", "").strip()
        quantity_row = sku_row + 1
        quantity = None
        while quantity_row <= sheet.max_row:
            try:
                quantity = int(sheet.cell(row=quantity_row, column=1).value)
                break
            except (ValueError, TypeError):
                quantity_row += 1

            # 如果当前行是最后一行，寻找下一个有效的数字作为数量
            if quantity_row == sheet.max_row and quantity is None:
                search_row = sheet.max_row
                while search_row > 0:
                    try:
                        quantity = int(sheet.cell(row=search_row, column=1).value)
                        break
                    except (ValueError, TypeError):
                        search_row -= 1

        weight_row = row + 1
        dimensions_row = row + 1
        while weight_row < sheet.max_row:
            weight_value = str(sheet.cell(row=weight_row, column=1).value)
            if "weight:" in weight_value.lower():
                weight = sheet.cell(row=weight_row + 1, column=1).value
                unit = sheet.cell(row=weight_row + 2, column=1).value
                break
            weight_row += 1

        while dimensions_row < sheet.max_row:
            dimensions_value = str(sheet.cell(row=dimensions_row, column=1).value)
            if "dimensions:" in dimensions_value.lower():
                length = sheet.cell(row=dimensions_row + 1, column=1).value
                width = sheet.cell(row=dimensions_row + 3, column=1).value
                height = sheet.cell(row=dimensions_row + 5, column=1).value
                unit_dimensions = sheet.cell(row=dimensions_row + 6, column=1).value
                break
            dimensions_row += 1

        shipment_data["Shipment ID"].append(shipment_id)
        shipment_data["Box ID"].append(box_id)
        shipment_data["SKU"].append(sku)
        shipment_data["数量"].append(quantity)
        shipment_data["Weight"].append(weight)
        shipment_data["Unit"].append(unit)
        shipment_data["Length"].append(length)
        shipment_data["Width"].append(width)
        shipment_data["Height"].append(height)
        shipment_data["UnitDimensions"].append(unit_dimensions)

        # 更新sku_row以跳过已处理的数量行
        sku_row = quantity_row

    elif "Box ID:" in sku_value:
        box_id = sku_value.replace("Box ID:", "").strip()

    return sku_row, box_id


# 读取Excel文件
filename = r'D:\\运营\\1数据源\\invoice\\shipment.xlsx'
wb = load_workbook(filename)

# 初始化变量和结果dataframe
shipment_data = {
    "Shipment ID": [],
    "Box ID": [],
    "SKU": [],
    "数量": [],
    "Weight": [],
    "Unit": [],
    "Length": [],
    "Width": [],
    "Height": [],
    "UnitDimensions": []
}


# 处理每个工作表
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    shipment_id = ""
    box_id = ""

    # 遍历第一列
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value

        if cell_value is not None:
            cell_value = str(cell_value)  # 将cell_value转换为字符串

            if "Shipment ID:" in cell_value:
                shipment_id = cell_value.replace("Shipment ID:", "").strip()

            elif "Box ID:" in cell_value:
                box_id = cell_value.replace("Box ID:", "").strip()

                # 处理SKU和数量
                sku_row = row + 1
                while sku_row <= sheet.max_row:
                    sku_row, box_id = process_sheet(sheet, sku_row, shipment_id, box_id, shipment_data)

                    # 如果当前行是最后一行，跳出循环
                    if sku_row == sheet.max_row:
                        break

                # 如果当前行是最后一行，跳出循环
                if row == sheet.max_row:
                    break

    # 将处理完的数据输出到invoicedraft.xlsx文件的Sheet1中
    shipment_df = pd.DataFrame(shipment_data)
    with pd.ExcelWriter('D:\\运营\\invoicedraft.xlsx', mode='a', engine='openpyxl') as writer:
        shipment_df.to_excel(writer, sheet_name='Sheet1', index=False)
    # 重置shipment_data
    shipment_data = {
        "Shipment ID": [],
        "Box ID": [],
        "SKU": [],
        "数量": [],
        "Weight": [],
        "Unit": [],
        "Length": [],
        "Width": [],
        "Height": [],
        "UnitDimensions": []
    }

print("程序执行完毕")


