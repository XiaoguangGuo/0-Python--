#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。
#此程序判断近期10周的数据，合并到一起。这样数据集不至于过大。
# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import shutil

print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中。此操作将更新周广告汇总表为最新10周以内")
a=input('按任意键继续，如未准备好则取消程序执行', )

newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)

#创建汇总表
bulkhzWorkbook=Workbook()
sheet=bulkhzWorkbook.active
Listcol=[ 'Record ID', 'Record Type','Campaign ID','Campaign','Campaign Daily Budget','Portfolio ID','Campaign Start Date','Campaign End Date','Campaign Targeting Type','Ad Group','Max Bid', 'Keyword or Product Targeting', 'Product Targeting ID', 'Match Type','SKU','Campaign Status','Ad Group Status','Status','Impressions','Clicks',
'Spend','Orders','Total Units', 'Sales' ,'ACoS','Bidding strategy','Placement Type','Increase bids by placement','Country','日期','周数']
sheet.append(Listcol)
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表NEW.xlsx')


#bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表2.xlsx')

sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
#print(sheet.max_row,sheet.max_column)

# 从第sheet.max_row+1行将表格New写入。
# 在表格的第  列写入国家名 第  列 写入日期  第  列写入 周数 。

#newdata=sheet

# 用pandas
#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'
for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
#判断是否是10周内的数据，如果是就读数据，如果不是，则显示此文件不读。
    if delta<11:
            sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
#Newdata=pd_read_excel(r'D:\运营\周bulk广告数据\D:/运营/周Bulk广告数据/GV-CA_bulk-a2ijxz1fd77inr-20210411-20210417-1620645025514.xlsx',row=1,
    #newdata=pd.read_excel(newfilepath+'GV-CA_bulk-a2ijxz1fd77inr-20210411-20210417-1620645025514.xlsx',engine="openpyxl",sheet_name="Sponsored Products Campaigns").assign(Country=os.path.basename(newdataname).split('_')[0], 日期=os.path.basename(newdataname).split('-')[4]) 
#增加周数并赋值1
            sourcedata['周数']=1
            #print(sourcedata)
            for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
                sheet.append(row)
                #print(sheet.max_row,sheet.max_column)
    else:
         print('不导入')
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')

#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。
    #sheet.cell(row=2, column=2).value = "omg"
    #print(sheet.cell(2,2).value)


for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=30).value
    print(type(b),b)
    #sheet.cell(row=i,column=30)=datetime.datetime.fromisoformat('2020-12-12 12:22:22') 这个值得试试。
    #sheet.cell(row=i,column=30)==datetime.datetime.strptime(b,'%Y%m%d')
    sheet.cell(row=i,column=30).value=datetime.datetime.strptime(b,'%Y%m%d')
    sheet.cell(row=i,column=30).number_format = 'yyyy-mm-dd hh:mm:ss'
    bnewtime=sheet.cell(row=i,column=30).value
    print(type(bnewtime),bnewtime)
    #bt=datetime.datetime.strptime(b,'%Y-%m-%d %H:%M:%S')
    #print(bt)
    c=(maxtime-bnewtime).days//7+1 #c=(a-b).days()错误
    print(maxtime,b,c)     
        #caozuoriqi= datetime.datetime.strptime(sheet.cell(row=i, column=30).value, "%Y%m/%d/")
        #print(caozuoriqi)
    #zhoushu=(('2021-10-30').dt.days-(cell(row=i,column=30).value).dt.days)//7+1
    sheet.cell(row=i, column=31).value =c
    
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')


