import pandas as pd
import openpyxl

import openpyxl

def find_blank_area(file_path, extra_rows=50):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    blank_area_start = -1

    consecutive_blank_rows = 0
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + extra_rows + 1):
        empty_cells = sum([1 for col in range(1, max_col + 1) if ws.cell(row=row, column=col).value is None])

        if empty_cells >= 6:
            consecutive_blank_rows += 1
        else:
            consecutive_blank_rows = 0

        if consecutive_blank_rows == 10:
            blank_area_start = row - 9
            break

    if blank_area_start > 0:
        col_names = [ws.cell(row=blank_area_start - 1, column=col).value for col in range(1, max_col + 1)]
        print(col_names)
        return col_names
    else:
        return None

def match_col_names(template_col_names, mapping_file_path, mapping_sheet_name):
    mapping_df = pd.read_excel(mapping_file_path, sheet_name=mapping_sheet_name)
    middle_col_names = []

    for col_name in template_col_names:
        middle_col_name = mapping_df.loc[mapping_df['模板列名'] == col_name, '中间列名']
        if not middle_col_name.empty:
            middle_col_names.append(middle_col_name.iloc[0])
        else:
            middle_col_names.append(None)

    return middle_col_names

if __name__ == "__main__":
    file_path = "D:\\运营\\发票模板\\template.xlsx"
    mapping_file_path = "D:\\运营\\invoice\\发票基础信息表.xlsx"
    mapping_sheet_name = "列名对应"
    
    template_col_names = find_blank_area(file_path)
    if template_col_names:
        middle_col_names = match_col_names(template_col_names, mapping_file_path, mapping_sheet_name)
        print("转换后的中间列名列表：", middle_col_names)
    else:
        print("在工作簿中未找到满足条件的空白区域。")


other_df=pd.read_excel(r'G:\\运营\invoice\\invoicedraft2.xlsx')
other_df_selected=other_df[middle_col_names]

other_df_selected.to_excel(r'G:\\运营\invoice\\invoicedraft-new.xlsx')
