import pandas as pd
from openpyxl import load_workbook


# 读取Excel文件
filename = r'D:\\运营\\1数据源\\invoice\\shipment.xlsx'
wb = load_workbook(filename)
sheet = wb.active

# 初始化变量和结果dataframe
shipment_data = {
    "Shipment ID": [],
    "Box ID": [],
    "SKU": [],
    "数量": []
}
shipment_id = ""
box_id = ""

# 遍历第一列
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    if cell_value is not None:
        cell_value = str(cell_value)  # 将cell_value转换为字符串

        if "Shipment ID:" in cell_value:
            shipment_id = cell_value.replace("Shipment ID:", "").strip()

        elif "Box ID:" in cell_value:
            box_id = cell_value.replace("Box ID:", "").strip()

            # 寻找SKU和Print
            sku_row = row + 1
            while sku_row < sheet.max_row:
                sku_value = str(sheet.cell(row=sku_row, column=1).value)
                if "SKU:" in sku_value:
                    sku = sku_value.replace("SKU:", "").strip()
                    print_row = sku_row + 1
                    while print_row < sheet.max_row:
                        print_value = str(sheet.cell(row=print_row, column=1).value)
                        if "Print" in print_value:
                            quantity = sheet.cell(row=print_row + 1, column=1).value
                            break
                        print_row += 1

                    shipment_data["Shipment ID"].append(shipment_id)
                    shipment_data["Box ID"].append(box_id)
                    shipment_data["SKU"].append(sku)
                    shipment_data["数量"].append(quantity)

                    sku_row = print_row  # 更新sku_row以跳过已处理的Print行
                elif "Box ID:" in sku_value:
                    break  # 跳出循环以处理下一个Box ID
                else:
                    sku_row += 1

shipment_df = pd.DataFrame(shipment_data)
print(shipment_df)
