# 使用Oenyxl编写，没有使用pandas
### 第一部分汇总各国bulk广告数据到汇总表
#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。

# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表


print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )

newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)


#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\2生成过程表\周bulk广告数据汇总表.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#？？？旧的程序，不要了bulkoperationworkbook=load_workbook(r'D:\运营\bulkoperation模板.xlsx')
#???sheetbulkoperation=bulkoperationworkbook.worksheets[0]

#指定来源文件
bulkdatafilepath = 'D:\\运营\\1数据源\\周bulk广告数据\\'


for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])

    columnlist=["Campaign Daily Budget","Max Bid","Spend","Sales" ]
#将逗号变成点
    sourcedata[columnlist]=sourcedata[columnlist].replace(',','.',regex=True).astype(float)
    sourcedata["ACoS"]=sourcedata["ACoS"].replace(',','.',regex=True)
    
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=1
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbook.save(r'D:\运营\2生成过程表\周bulk广告数据汇总表.xlsx') #汇总所有广告数据

 
#???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

#拷贝一份sourcedata到bulkoperation文件夹,文件名不变
    shutil.copy(r'D:\\运营\\1数据源\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\\运营\\1数据源\\bulkoperationfiles\\')
#移动广告数据到历史数据 
    shutil.copy(r'D:\\运营\\1数据源\\周bulk广告数据\\'+ str(bulkdatafile),r'D:\\运营\\HistoricalData\\周bulk广告数据\\')
    
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。
 
Allbulkpath='D:\\运营\\2生成过程表\\'
writer=pd.ExcelWriter(Allbulkpath+'周bulk数据Summary.xlsx')
Allbulk=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx')
Allbulk["周数"]=(maxtime-Allbulk["日期"]).dt.days//7+1                      
Allbulk.to_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx',sheet_name="Sheet",index=False)  
 
###???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

#######################以下生成Summary的程序##################################################################***************

from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil

#在这之前要生成汇总表，并且把每个国家的bulk表备份到D:\\运营\\bulkoperationfiles\\


 
writer=pd.ExcelWriter(Allbulkpath+'周bulk数据Summary.xlsx')
 
Allbulkold1=pd.read_excel(r'D:\\运营\\1数据源\\周Bulk广告数据汇总表历史\\'+"周bulk广告数据汇总表_2022-8-27_2022-9-24.xlsx")
Allbulkold2=pd.read_excel(r'D:\\运营\\1数据源\\周Bulk广告数据汇总表历史\\'+"周bulk广告数据汇总表_2022-5-28_2022-8-20.xlsx")



Allbulk=pd.concat([Allbulk,Allbulkold1,Allbulkold2])
Allbulk["周数"]=(maxtime-Allbulk["日期"]).dt.days//7+1


print("以下生成Summary的程序")




#定义bulk数据汇总表所在路径
Allbulkpath='D:\\运营\\2生成过程表\\'
Allbulk=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx',dtype = {'SKU':str})



AllbulkCampaign=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign","Campaign Targeting Type"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaign["zhuanhualv"]=AllbulkCampaign['Orders']/AllbulkCampaign['Clicks']
AllbulkCampaign_List=AllbulkCampaign["Campaign"].drop_duplicates().to_list()
AllbulkCampaign_Country_List=AllbulkCampaign["Country"].drop_duplicates().to_list()

#AllbulkCampaign1week=Allbulk[(Allbulk['Record Type']=="Campaign")&(Allbulk['周数']==1)].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')


AllbulkCampaignWEEK=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')

AllbulkSKUCampaignWEEK_1=Allbulk[(Allbulk['Record Type']=="Ad")&(Allbulk['周数']==1)].groupby(["Country","Campaign","SKU","Ad Group"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkSKUCampaignWEEK_1["zhuanhualv"]=AllbulkSKUCampaignWEEK_1["Orders"]/AllbulkSKUCampaignWEEK_1["Clicks"]

AllbulkCampaignSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","Campaign","周数","Campaign Status"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignSKUTotal=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","Campaign","Ad Group"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignSKUWEEK["Campaign Targeting Type"]=""
AllbulkCampaignSKUTotal["Campaign Targeting Type"]=""

###########################################################################################################

for AllbulkCampaign_Country_oi99 in AllbulkCampaign_Country_List:
    AllbulkCampaign_Country99=AllbulkCampaign.loc[AllbulkCampaign["Country"]==AllbulkCampaign_Country_oi99]
    AllbulkCampaign_Country99_CampaignList= AllbulkCampaign_Country99.loc[AllbulkCampaign_Country99["Country"]==AllbulkCampaign_Country_oi99,"Campaign"].drop_duplicates().to_list()
    
    for campaign_oi99 in AllbulkCampaign_Country99_CampaignList:
        campaigntype99=AllbulkCampaign_Country99.loc[(AllbulkCampaign_Country99["Country"]==AllbulkCampaign_Country_oi99)&(AllbulkCampaign_Country99["Campaign"]==campaign_oi99),"Campaign Targeting Type"].values[0] 
        AllbulkCampaignSKUTotal.loc[(AllbulkCampaignSKUTotal["Campaign"]==campaign_oi99)&(AllbulkCampaignSKUTotal["Country"]==AllbulkCampaign_Country_oi99),"Campaign Targeting Type"]=campaigntype99
############################################################################################################################campaigntype99


###########################################################################################################
 


AllbulkCampaignSKUTotal["zhuanhualv"]=AllbulkCampaignSKUTotal["Orders"]/AllbulkCampaignSKUTotal["Clicks"]

AllbulkCampaignSKUTotal["zhuanhualv_rank1"]=AllbulkCampaignSKUTotal.groupby(["Country","SKU","Campaign Targeting Type"],as_index=False)[["zhuanhualv"]].rank(ascending=0,method='max')
AllbulkCampaignSKUTotal["zhuanhualv_rank2"]=AllbulkCampaignSKUTotal.groupby(["Country","SKU","Campaign Targeting Type"],as_index=False)[["zhuanhualv"]].rank(ascending=0,method='dense')                                                                                       

AllbulkCampaignSKUTotalzhuanhualvMax=AllbulkCampaignSKUTotal.groupby(["Country","Campaign","Ad Group","SKU","Campaign Targeting Type"],as_index=False)[["zhuanhualv"]].agg('max')

AllbulkCampaignSKUWEEK["Campaign Targeting Type"]=""

########################################################################################################################################33                                               
for AllbulkCampaign_Country_oi in AllbulkCampaign_Country_List:
    
    AllbulkCampaign_Country88=AllbulkCampaign.loc[AllbulkCampaign["Country"]==AllbulkCampaign_Country_oi]
    AllbulkCampaign_Country_CampaignList88= AllbulkCampaign_Country88.loc[AllbulkCampaign_Country88["Country"]==AllbulkCampaign_Country_oi99,"Campaign"].drop_duplicates().to_list()

    
    for campaign_oi in AllbulkCampaign_Country_CampaignList88:
        campaigntype88=AllbulkCampaign_Country88.loc[(AllbulkCampaign_Country88["Country"]==AllbulkCampaign_Country_oi)&(AllbulkCampaign_Country88["Campaign"]==campaign_oi),"Campaign Targeting Type"].values[0]
        AllbulkCampaignSKUWEEK.loc[(AllbulkCampaignSKUWEEK["Campaign"]==campaign_oi)&(AllbulkCampaignSKUWEEK["Country"]==AllbulkCampaign_Country_oi),"Campaign Targeting Type"]=campaigntype88





                                               
AllbulkCampaignSKUWEEK["zhuanhualv"]=AllbulkCampaignSKUWEEK["Orders"]/AllbulkCampaignSKUWEEK["Clicks"]

AllbulkSKUMax=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","Campaign","Ad Group","SKU"],as_index=False)[['Spend']].max()


                                                                              
AllbulkCampaignKeywordWEEK=Allbulk.groupby(["Country","Campaign","Keyword or Product Targeting","Ad Group","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")




#以下为建立Camaign和SKU联系的程序
ALLbulkCampaignSKU=Allbulk[['Country','Campaign','SKU']]

ALLbulkCampaignSKU=ALLbulkCampaignSKU.drop_duplicates()
ALLbulkCampaignSKU=ALLbulkCampaignSKU.dropna(axis=0,how='any')
print(ALLbulkCampaignSKU)

CamaignSKUAgg=ALLbulkCampaignSKU.groupby(["Country","Campaign"],as_index=False).agg({'SKU':[",".join]})#追加的新的汇总comaignSKU
New_columns=['Country',"Campaign",'SKU']


AllbulkSKUMax_CampaignWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","Campaign","SKU","周数","Campaign Status"],as_index=False)[['Spend']].agg('max')
AllbulkSKU_Campaign=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","Campaign","SKU"],as_index=False)[['Spend','Orders','Clicks']].agg('sum').reset_index()
AllbulkSKU_Campaign["zhuanhualv"]=AllbulkSKU_Campaign["Orders"]/AllbulkSKU_Campaign["Clicks"]
AllbulkSKU_Campaign["SKU-Campaign-zhuanhualv-ranking"]=AllbulkSKU_Campaign.groupby(["Country","SKU"],as_index=False)[['zhuanhualv']].rank(ascending=0,method='max')

AllbulkSKU_Campaign["Campaign-SKU_Spend_ranking"]=AllbulkSKU_Campaign.groupby(["Country","Campaign"],as_index=False)[['Spend']].rank(ascending=0,method='max')
AllbulkSKU_Campaign["SKU_Campaign_Spend_ranking"]=AllbulkSKU_Campaign.groupby(["Country","SKU"],as_index=False)[['Spend']].rank(ascending=0,method='max')

AllbulkSKU_Campaignrank1=AllbulkSKU_Campaign[AllbulkSKU_Campaign["Campaign-SKU_Spend_ranking"]==1]
AllbulkCampaignSKUWEEK["Spend_Order"]=AllbulkCampaignSKUWEEK.groupby(["Country","Campaign","SKU","周数","Campaign Status","Campaign Targeting Type"],as_index=False)[['Spend']].rank(ascending=0,method='max')
AllbulkCampaignSKUWEEK["Click_Order"]=AllbulkCampaignSKUWEEK.groupby(["Country","Campaign","SKU","周数","Campaign Status","Campaign Targeting Type"],as_index=False)[['Clicks']].rank(ascending=0,method='max')
#SKUCampaign_zhouzhuqanlv_Max

                                                                              
CamaignSKUAgg.columns=New_columns




writer=pd.ExcelWriter(Allbulkpath+'周bulk数据Summary.xlsx')
AllbulkSKUCampaignWEEK_1.to_excel(writer,"SKUCampaignWEEK_1")
AllbulkCampaign.to_excel(writer,"Campaign汇总")
AllbulkCampaignWEEK.to_excel(writer,"CampaignWEEK汇总")
AllbulkSKUWEEK.to_excel(writer,"SKU-WEEK")
AllbulkCampaignSKUWEEK.to_excel(writer,"SKU-Campaign-WEEK",index=False)
AllbulkCampaignKeywordWEEK.to_excel(writer,"Keyword-Campaign-WEEK")
AllbulkSKU_Campaign.to_excel(writer,"SKU-Campaign-Spend")
AllbulkSKU_Campaignrank1.to_excel(writer,"SKUMax-Campaign")
AllbulkSKUMax_CampaignWEEK.to_excel(writer,"SKUMax-Campaign-WEEK") 
AllbulkCampaignSKUTotal.to_excel(writer,"AllSKUCampaign")
CamaignSKUAgg.to_excel(writer,"CamaignSKUAgg")#追加的新的汇总comaignSKU
AllbulkCampaignSKUTotalzhuanhualvMax.to_excel(writer,"CampaignSKUTotalzhuanhualvMax",index=False)

writer.close()


######################################以下为做Biaotou周汇总###################################################



######################################以下为做Biaotou周汇总###################################################


# -*- coding:utf-8 –*-
import os
import pandas as pd
import shutil

# -*- coding:utf-8 –*-
import os
import pandas as pd
import shutil
import datetime

CampaignSKU_Summary=pd.read_excel(r'D:/运营/2生成过程表/周bulk数据Summary.xlsx',sheet_name="SKU-Campaign-WEEK")

CampaignSKU_SummarySum=CampaignSKU_Summary.groupby(["Country","SKU","Campaign"],as_index=False)[["Impressions","Clicks","Spend","Orders","Total Units","Sales"]].agg('sum')



CampaignSKU_Summary["皮质层标签"]=" "

CampaignSKU_Summary["Zhouzhuanlv"]=CampaignSKU_Summary["Orders"]/CampaignSKU_Summary["Clicks"]

CampaignSKU_SummarySum["Zhouzhuanlv"]=CampaignSKU_SummarySum["Orders"]/CampaignSKU_SummarySum["Clicks"]


CampaignSKU_Summary.loc[(CampaignSKU_Summary["Clicks"]>0) &(CampaignSKU_Summary["Zhouzhuanlv"]>0.15),"皮质层标签"] = CampaignSKU_Summary["皮质层标签"].astype(str)+"好广告"


CampaignSKU_Summary.loc[(CampaignSKU_Summary["Clicks"]>0) &(CampaignSKU_Summary["Zhouzhuanlv"]<0.05),"皮质层标签"] = CampaignSKU_Summary["皮质层标签"].astype(str)+"差广告"

#CampaignSKU_Summary10=CampaignSKU_Summary.loc[(CampaignSKU_Summary["周数"]<5)&(CampaignSKU_Summary["Country"]=="GV-US")]

CampaignSKU_Summary_biaotou=CampaignSKU_Summary[["Country","SKU","Campaign"]].drop_duplicates()
CampaignSKU_Summary_biaotou=pd.merge(CampaignSKU_Summary_biaotou,CampaignSKU_SummarySum,on=["Country","SKU","Campaign"] ,how="left")
print(CampaignSKU_Summary_biaotou)

for i in range(1,20):
    #CampaignSKU_Summary_i=CampaignSKU_Summary["Clicks","Orders"].loc[(CampaignSKU_Summary["周数"]==i)]
    CampaignSKU_Summary_i=CampaignSKU_Summary.loc[(CampaignSKU_Summary["周数"]==i)]
    
    #CampaignSKU_Summary_i=CampaignSKU_Summary_i["Country","SKU","Campaign","Clicks","Orders"]
    #更改列名

    CampaignSKU_Summary_i.rename(columns = {'Clicks':'Clicks'+str(i), 'Orders':'Orders'+str(i),'Spend':'Spend'+str(i),'Impressions':'Impressions'+str(i)}, inplace = True)

    CampaignSKU_Summary_biaotou=pd.merge(CampaignSKU_Summary_biaotou,CampaignSKU_Summary_i,on=["Country","SKU","Campaign"] ,how="left")
    




#CampaignSKU_Summary_pivot10=CampaignSKU_Summary10.pivot_table(values=["Clicks","Orders"], index=['Country','SKU','Campaign'],columns="周数", aggfunc = 'sum', fill_value=None, margins=False, dropna=False,margins_name='All').reset_index() # 是否启用总计行/列# 值

print(CampaignSKU_Summary_biaotou)


CampaignSKU_Summary_biaotou.to_excel(r'D:\\运营\\2生成过程表\\CampaignSKU_Summary_biaotou.xlsx',sheet_name="sheet1",startrow=0,header=True,index=True)




