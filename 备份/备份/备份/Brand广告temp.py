
# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表

print("请确认老站计划数据已经就位")
print("请确认新站计划数据已经就位")
print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )


newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)


#导入历史汇总表（可以汇总历史所有情况）
#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("品牌表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
 
    #???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

    
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。


for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=39).value

    bnewtime=sheet.cell(row=i,column=39).value

    c=(maxtime-bnewtime).days//7+1

    sheet.cell(row=i, column=40).value =c
    
                      
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')
