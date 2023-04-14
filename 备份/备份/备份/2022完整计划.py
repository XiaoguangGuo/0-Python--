
# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表

print("请确认老站计划数据已经就位")
print("请确认新站计划数据已经就位")
print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )


newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)



# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil



#导入历史汇总表-品牌（可以汇总历史所有情况）

bulkhzWorkbookbrand=load_workbook(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')


sheetnames=bulkhzWorkbookbrand.sheetnames
print("品牌表名",sheetnames)
print(bulkhzWorkbookbrand.sheetnames)
sheet=bulkhzWorkbookbrand[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'

for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name="Sponsored Brands Campaigns").assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
    sourcedata=sourcedata.dropna(axis=0, how='all', thresh=None, subset=None, inplace=False)#只有全部为空才回被删除
    
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=delta
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbookbrand.save(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')#汇总所有广告数据

 




#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#？？？旧的程序，不要了bulkoperationworkbook=load_workbook(r'D:\运营\bulkoperation模板.xlsx')
#???sheetbulkoperation=bulkoperationworkbook.worksheets[0]

#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'

for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
    sourcedata.replace(",",".",inplace=True)
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=1
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')#汇总所有广告数据

 
    #???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

#拷贝一份sourcedata到bulkoperation文件夹,文件名不变
    shutil.copy(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\\运营\\bulkoperationfiles\\')
#移动广告数据到历史数据 
    shutil.move(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile),r'D:\\运营\\HistoricalData\\周bulk广告数据\\')
    
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。


for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=30).value

    bnewtime=sheet.cell(row=i,column=30).value

    c=(maxtime-bnewtime).days//7+1

    sheet.cell(row=i, column=31).value =c
    
                      
    
#保存所有广告数据汇总表
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')
###???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')
########################################################################################################################################################################
#######################运行汇总Summary

#使用Oenyxl编写，没有使用pandas
### 第一部分汇总各国bulk广告数据到汇总表
#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。

# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil



#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

 
print("以下为处理bulk操作报表的程序")

##先用临时文件测试后copy来
#此程序测试完后应该考入bulkoperation程序中

print("以下为处理bulk操作报表的程序")



#定义bulk数据汇总表所在路path='D:\\运营\\'
Allbulkpath='D:\\运营\\'

Allbulk=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx')
AllbulkBrand=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表-品牌.xlsx')
ListingPd=pd.read_excel(r'D:/2019plan/Listing.xlsx')

 

#####生成各种Campaign summary

 

AllbulkCampaign=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
#AllbulkCampaign1week=Allbulk[(Allbulk['Record Type']=="Campaign")&(Allbulk['周数']==1)].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkFenlei=pd.merge(Allbulk,ListingPd,how="left",left_on=["Country","SKU"],right_on=["COUNTRY","seller-sku"])

AllbulkCampaignWEEK=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkFenleiWeek=AllbulkFenlei.groupby(["Country","大类","小类","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignSKUWEEK_fenlei=pd.merge(AllbulkCampaignSKUWEEK,ListingPd,how="left",left_on=["Country","SKU"],right_on=["COUNTRY","seller-sku"])

AllbulkCampaignKeywordWEEK=Allbulk.groupby(["Country","Campaign","Keyword or Product Targeting","Ad Group","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")
AllbulkCampaignKeywordTOTAL=Allbulk.groupby(["Country","Campaign","Keyword or Product Targeting","Ad Group" ],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")

AllbulkCOUNTRYSKUCAMPAIGN=Allbulk.groupby(["Country","Campaign","SKU" ],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")


#品牌

AllbulkBrandCampaignWEEK=AllbulkBrand[AllbulkBrand['Record Type']=="Campaign"].groupby(["Country","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkBrandCampaignKeywordWEEK=AllbulkBrand.groupby(["Country","Campaign","Keyword","Ad Group","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")




#写入Excel表格
writer=pd.ExcelWriter(Allbulkpath+'周bulk数据Summary.xlsx')
AllbulkCampaign.to_excel(writer,"Campaign汇总")
AllbulkCampaignWEEK.to_excel(writer,"CampaignWEEK汇总")
AllbulkSKUWEEK.to_excel(writer,"SKU-WEEK")
AllbulkCampaignSKUWEEK.to_excel(writer,"SKU-Campaign-WEEK")
AllbulkFenleiWeek.to_excel(writer,"AllbulkFenleiWeek")
AllbulkCampaignKeywordWEEK.to_excel(writer,"Keyword-Campaign-WEEK")
AllbulkCampaignKeywordTOTAL.to_excel(writer,"CAMPAIGN-KEYEORD-TOTAL")
AllbulkCOUNTRYSKUCAMPAIGN.to_excel(writer,"COUNTRY-CAMPAIGN-SKU")
AllbulkCampaignSKUWEEK_fenlei.to_excel(writer,"COUNTRY-CAMPAIGN-SKU_fenlei")

AllbulkBrandCampaignWEEK.to_excel(writer,"Brand-Campaign-week")
AllbulkBrandCampaignKeywordWEEK.to_excel(writer,"Brand-Campaign-keyword-week")




writer.save()



#####以下为做计划
########################################################################################################################################################################
####检查文件有效性

# -*- coding:utf-8 –*-
import os
import pandas as pd
import time



#学习素材
#遍历数据文件
#sourcedata=pd.read_csv(salesfilepath+str(salesfile)).assign(Country=os.path.basename(salesfile).split('_')[0], 日期=os.path.basename(salesfile).split('_')[1])
#sourcedata['日期'] = pd.to_datetime(sourcedata['日期'])
#print(sourcedata[['Country','日期']])
#for i in range(len(key)):
    #遍历国家名     

def is_valid_date1(strdate):
  #'''判断是否是一个有效的日期字符串'''
  try:
    (time.strptime(strdate, "%Y-%m-%d")) 
    return True
  except:
    return False
def is_valid_date2(strdate):
  #'''判断是否是一个有效的日期字符串'''
  try:
    (time.strptime(strdate, "%Y%m%d")) 
    return True
  except:
    return False

#计划数据的路径
Plandatapath=r'D:\\运营\\2019lan'

CheckPathlist=[r'D:/运营/计划数据/老站/',r'D:/运营/计划数据/Newcountries']

Filecliplist=["销售数据","当日库存","在途库存"]
Granvelalist=["US","CA","MX"]
GranvelaGVList=["GV-US","GV-CA","GV-MX"]
NewStoreslist=["NEW-CA","NEW-US","NEW-MX","NEW-JP","NEW-UK","NEW-ES","NEW-IT","NEW-FR","NEW-NL","HM-US"]


# 名称都在清单里，名称不重复，国家数量缺少的报出来，漏掉国家了没？确认。
#日期格式是否正确
#检查文件列数和列名
#检查各个国家是否有10周的数据。
#检查在途库存的到货日期


for i in CheckPathlist:
    
    print(i)
    print("----"*10)
    print("----"*10)
    for j in Filecliplist:
        countrynamelist=[]
        
        print(j)
        print("--"*10)
#for file in os.listdir(src_dir_path_inventory):

    
    #print(os.listdir(src_dir_path_inventory))

# 文件名列表
        checkfilelist=os.listdir(i+j)

        print(checkfilelist,"文件个数",len(checkfilelist))
        if ("老站" in i) and (len(checkfilelist)==3):
            print("老站"+j+"文件数量正确")
        elif("新站" in i) and (len(checkfilelist)==9):
            print("新站"+j+"文件数量正确")
            
        else:
           
            print("!!!!!!!"*5)
            print(i+j+"  文件数量不正确")
        
        for k in checkfilelist:
            countryname=os.path.basename(k).split('_')[0]
            countrynamelist.append(countryname)
            if i==r'D:/运营/计划数据/Newcountries':
                if countryname in Granvelalist:
                    print (countryname,"OK")
                else:
                    print("!!!!!!!"*5)
                    print(i,"国家名不对")
                
            elif i==r'D:/运营/计划数据/新站/':
                  if countryname in NewStoreslist:
                      print (countryname,"OK")
                  else:
                      print("!!!!!!!"*5)
                      print(j,"国家名不对",countryname)
                      
            if j=="销售数据":
                strdate=os.path.basename(k).split('_')[1]

                print(strdate,type(strdate))
                
                if is_valid_date1(strdate) or is_valid_date2(strdate):
                    
                    print( (is_valid_date1(strdate) or is_valid_date2(strdate) ))
                    
                    print("日期格式正确")
                else:
                    print("日期格式不对")
         
                
        if len(countrynamelist)!=len(set(countrynamelist)):
            print("!!!!!!!"*5)
            print("****注意注意注意****"+i+j+ "国家名有重复项！！")

        else:
            print(i+j+ "国家名无重复项")



import pandas as pd
import os
 


########################################################################################################################################################################
#检查并补充缺少的周数
pathlist=[r'D:/2019plan/周销售数据.xlsx',r'D:/SailingstarFBA计划/NEW-ALL周销售数据.xlsx',r'D:/2019plan/Mexico周销售数据.xlsx',r'D:/2019plan/Canada周销售数据.xlsx']



for j in pathlist:
    
    if j==r'D:/SailingstarFBA计划/NEW-ALL周销售数据.xlsx':
        salesdata=pd.read_excel(j,converters = {"周数": int})
        Countrylist=salesdata["Country"].unique().tolist()
        
        print(j)
        
        for k in Countrylist:
            
            saledataCountry=salesdata[(salesdata["Country"]==k)&(salesdata["周数"]<11)]
            print(saledataCountry)
            weeklist=saledataCountry["周数"].unique().tolist()

            print(weeklist)
            t=0
            for i in range(1,11):
                maxoriginal=len(salesdata)
                if i  not in weeklist:
                    print("!!!!!"*5)
                    print(k+"国家缺少"+str(i)+"周")
                    salesdata.loc[(maxoriginal+1),"Country"]=k
                    salesdata.loc[(maxoriginal+1),"SKU"]="test"
                    salesdata.loc[(maxoriginal+1),"周数"]=i
                    salesdata["周数"].round()
                    print(salesdata)
                    t=t+1
                    
                else:
                   
                    print("新站"+str(k)+str(i)+"周NEW-ALL周销售数据okokok")
               
    else:
    
        salesdata=pd.read_excel(j,converters = {"周数": int})
        salesdata2=salesdata[salesdata["周数"]<11]
        weeklist2=salesdata2["周数"].unique().tolist() 
        print(weeklist2)
        maxoriginal=len(salesdata)
        t=0
        for i in range(1,11):
           
            if i not in weeklist2:
                print(j+"国家缺少第"+i+"周")
                salesdata.loc[(maxoriginal+1),"SKU"]="test"
                salesdata.loc[(maxoriginal+1),"周数"]=i
                t=t+1
            else:
                print(j+"第"+str(i)+"周数不缺，ok")

    if t>0:
        salesdata.to_excel(j,index = False)
        print(j+"文件更新")
    else:
        print("无更新"+j)
                
            
        
input("请确认计划数据是否正确？如果正确则回车，如果需要更改，则先进行更改再回车。也可以中断程序.")

########################################################################################################################################################################

####运行老站计划
####运行老站计划
print("以下为运行老站计划")
print("以下为运行老站计划")
input("请确认计划数据是否正确？如果正确则回车，如果需要更改，则先进行更改再回车。也可以中断程序.")


# -*- coding:utf-8 –*-
import os
import pandas as pd
src_dir_path_inventory=r'D:\运营\计划数据\老站\当日库存'

key =['US','CA','MX']
t=key[0]
print(t)
#获取原来库存文件的列名
data_inventory_US=pd.read_excel(r'D:\2019plan\当日Amazon库存.xlsx')
data_inventory_CA=pd.read_excel(r'D:\2019plan\Canada当前Amazon库存.xlsx')
data_inventory_MX=pd.read_excel(r'D:\2019plan\Mexico当日Amazon库存.xlsx')
                                
inventorycolumns_US=data_inventory_US.columns.tolist()
inventorycolumns_CA=data_inventory_CA.columns.tolist()
inventorycolumns_MX=data_inventory_MX.columns.tolist()
                                
print(inventorycolumns_US)

# 在文件夹里查找文件

for file in os.listdir(src_dir_path_inventory):
    print(os.listdir(src_dir_path_inventory))
    
    data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+ str(file),encoding="Latin1")    # 读取以encoding='Latin1'分        
    if key[0] in file:
        print(file)
    # 执行语句
        print("有US库存")
       
         # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
                                                 
        data_csv.columns=inventorycolumns_US                      
        data_csv.to_excel(r'D:\2019plan\当日Amazon库存.xlsx',sheet_name="当前Amazon库存",startrow=0,header=True,index=False)
        
    elif key[1]in file:
        print("有CA库存")
        print(file)
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        data_csv.columns=inventorycolumns_CA    
                         
        data_csv.to_excel(r'D:\2019plan\Canada当前Amazon库存.xlsx', sheet_name="15828640259018099",startrow=0,header=True,index=False)

        
      
    elif key[2]in file:
        print("有MX库存")
        print(file)        
        #df_data.columns.tolist())
        
        data_csv.columns=inventorycolumns_MX  
                         
        data_csv.to_excel(r'D:\2019plan\Mexico当日Amazon库存.xlsx', sheet_name="当前Amazon库存",startrow=0,header=True,index=False)

        print(data_csv)
        
    else:
        print("什么库存文件都没有")

# 导入reStock

src_dir_path_restock=r'D:\运营\计划数据\老站\restock'
print(os.listdir(src_dir_path_restock))
for file in os.listdir(src_dir_path_restock):
    data_csv2 = pd.read_table(r'D:\\运营\\计划数据\\老站\\restock\\'+ str(file),encoding="Latin1")    # 读取以分        
    if key[0] in file:
        print(file)
    # 执行语句
        print("有USrestock")
       
         # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        print(data_csv2)
                     
        data_csv2.to_excel(r'D:\2019plan\restock-report.xlsx',sheet_name="restock-report",startrow=0,header=True,index=False)
        
    elif key[1]in file:
        print("有CArestock")
        print(file)
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        
                         
        data_csv2.to_excel(r'D:\2019plan\restock-report_CA.xlsx', sheet_name="REstock-CA",startrow=0,header=True,index=False)

        print(data_csv2)
      
    elif key[2]in file:
        print("有MXrestock")
        print(file)        
        #df_data.columns.tolist())
   
        data_csv2.to_excel(r'D:\2019plan\restock-report_MX.xlsx', sheet_name="restock-report_MX",startrow=0,header=True,index=False)

        print(data_csv2)
        
    else:
        print("什么restock文件都没有")

#复制销售数据 20210221模块待写入
src_dir_path_sales=r'D:\运营\计划数据\老站\销售数据'
# 设置来源文件搜索目录
print(os.listdir(src_dir_path_sales))
key =['US','CA','MX']
#设置需要搜索的国家名字

# 以后做函数来简化程序def data_csv_open(file)
# def sourcesales_totargetsales(path,listofcountry,target_excel)未来做


for file in os.listdir(src_dir_path_sales):
     
    data_sales_US=pd.read_excel(r'D:\2019plan\周销售数据.xlsx')
    data_sales_CA=pd.read_excel(r'D:\2019plan\Canada周销售数据.xlsx')
    data_sales_MX=pd.read_excel(r'D:\2019plan\Mexico周销售数据.xlsx')
#未来可以做一个文件名列表包含文件名和sheet名
    salescolumns_US=data_sales_US.columns.tolist()
    salescolumns_CA=data_sales_CA.columns.tolist()
    salescolumns_MX=data_sales_MX.columns.tolist()
#取得目标文件的dataframe和列名

    if key[0] in file:
        print("开始处理US数据")
    
   
   
        data_csv_sales =pd.read_csv(r'D:\\运营\\计划数据\\老站\\销售数据\\'+ str(file)).assign(日期=os.path.basename(file).split('_')[1])
    #读取源数据加日期 把文件名中的日期写进来
        data_csv_sales['日期'] = pd.to_datetime(data_csv_sales['日期'])
        print(data_csv_sales['日期'])
        data_csv_sales['周数']=""
            
        print(data_csv_sales)
      
        ru=data_sales_US.columns.size-data_csv_sales.columns.size 
        
        if ru==0:
        #如果列数相同
            data_csv_sales.columns=salescolumns_US
            data_sales_US=data_sales_US.append(data_csv_sales,ignore_index=True)
        #做append将源数据合并到目标文件
            maxtime=pd.to_datetime(data_sales_US["日期"].max())
        #查目标文件的最晚日期
            print("最晚时间",maxtime)
            data_sales_US ['周数']=(maxtime-data_sales_US['日期']).dt.days//7+1
        #周数写到目标文件
        #在导出之前加周数
            data_sales_US.to_excel(r'D:\2019plan\周销售数据.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
            print("US销售数据更新完成")
        else:
            print("US销售数据未导出，请修改目标文件以保证列数相同")
            print("列数新下载数据文件和目标文件分别为：",data_csv_sales.columns.size,data_sales_CA.columns.size)
              

              
    # CA
    
    elif key[1] in file:
 
    
   
   
        data_csv_sales =pd.read_csv(r'D:\\运营\\计划数据\\老站\\销售数据\\'+ str(file),encoding="Latin1").assign(日期=os.path.basename(file).split('_')[1])
    #读取源数据加日期 把文件名中的日期写进来
        data_csv_sales['日期'] = pd.to_datetime(data_csv_sales['日期'])
        print(data_csv_sales['日期'])
        data_csv_sales['周数']=""
            
        print(data_csv_sales)
   
       
    
        
   
            
        ru=data_sales_CA.columns.size-data_csv_sales.columns.size 
        
        if ru==0:
        #如果列数相同
            data_csv_sales.columns=salescolumns_CA
            data_sales_CA=data_sales_CA.append(data_csv_sales,ignore_index=True)
        #做append将源数据合并到目标文件
            maxtime=pd.to_datetime(data_sales_CA["日期"].max())
        #查目标文件的最晚日期
            print("最晚时间",maxtime)
            data_sales_CA ['周数']=(maxtime-data_sales_CA['日期']).dt.days//7+1
        #周数写到目标文件
        #在导出之前加周数
            data_sales_CA.to_excel(r'D:\2019plan\Canada周销售数据.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
            print("CA销售数据更新完成")
        else:
            print("CA销售数据未导出，请修改目标文件以保证列数相同")
            print("列数新下载数据文件和目标文件分别为：",data_csv_sales.columns.size,data_sales_CA.columns.size)
              
    # MX

    elif key[2] in file:
        print("开始处理MX数据")
        # 不需要的 data_csv3 = pd.read_table(r'D:\\运营\\计划数据\\老站\\销售数据\\'+ str(file))
    # 打开原文件的dataframe
     

   
        data_csv_sales =pd.read_csv(r'D:\\运营\\计划数据\\老站\\销售数据\\'+ str(file)).assign(日期=os.path.basename(file).split('_')[1])
    #加日期把文件名中的日期写进来
        data_csv_sales['日期'] = pd.to_datetime(data_csv_sales['日期'])
        print(data_csv_sales['日期'])
       
        data_csv_sales['周数']=""
  
    #加周数
        ru=data_csv_sales.columns.size-data_sales_MX.columns.size
        if ru==0:
    #给列名赋值确保可以
            data_csv_sales.columns=salescolumns_MX
    #做append
            data_sales_MX=data_sales_MX.append(data_csv_sales,ignore_index=True)
            maxtime=pd.to_datetime(data_sales_MX["日期"].max())
            print(maxtime)
            data_sales_MX['周数']=(maxtime-data_sales_MX['日期']).dt.days//7+1
            data_sales_MX.to_excel(r'D:\2019plan\Mexico周销售数据.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
            print("MX销售数据更新完成")
        else:
            print("请修改目标文件，以保证列数相同")
            print("列数新下载数据文件和目标文件分别为：",data_csv_sales.columns.size,data_sales_MX.columns.size) 
    else:
        print("什么销售文件都没有")
    

    
    
# 复制TSV在途库存

src_dir_path_shipped=r'D:\运营\计划数据\老站\在途库存'
print(os.listdir(src_dir_path_shipped))


for file in os.listdir(src_dir_path_shipped):
    
    data_shipped_US=pd.read_excel(r'D:\2019plan\在途库存.xlsx')
    data_shipped_CA=pd.read_excel(r'D:\2019plan\Canada在途库存.xlsx')
    data_shipped_MX=pd.read_excel(r'D:\2019plan\Mexico在途库存.xlsx')
    salescolumns_US=data_shipped_US.columns.tolist()
    salescolumns_CA=data_shipped_CA.columns.tolist()
    salescolumns_MX=data_shipped_MX.columns.tolist()
      
    data_tsv5= pd.read_csv(r'D:\\运营\\计划数据\\老站\\在途库存\\'+ str(file),sep='\t',nrows =5)    
    batchnumber= data_tsv5.iat[0,1]
    data_tsv5= pd.read_csv(r'D:\\运营\\计划数据\\老站\\在途库存\\'+ str(file),sep='\t',header=6)    # 读取以分        
    data_tsv5["批次"]=batchnumber
  
    if key[0] in file:
        print(file)
    # 执行语句
        print("有US在途")

        data_tsv5['到货日期']=""
        data_tsv5['周数']=""
        print("lIESHU",data_tsv5.columns,salescolumns_US)
        data_tsv5.columns=salescolumns_US
        
    
        print(data_tsv5)
        data_shipped_US=data_shipped_US.append(data_tsv5,ignore_index=True)
                     
       #追加到在途计划 data_csv2.to_excel(r'D:\2019plan\restock-report.xlsx',sheet_name="restock-report",startrow=0,header=True,index=False)

        data_shipped_US.to_excel(r'D:\2019plan\在途库存.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
        print("US在途更新完成")
    #CA
    elif key[1]in file:
        print(file)
    # 执行语句
        print("有CA在途")   
         

        data_tsv5['到货日期']=""
        data_tsv5['周数']=""
        print("lIESHU",data_tsv5.columns,salescolumns_CA)
        data_tsv5.columns=salescolumns_CA
        
    
        print(data_tsv5)
        data_shipped_CA=data_shipped_CA.append(data_tsv5,ignore_index=True)
                     
       #追加到在途计划 data_csv2.to_excel(r'D:\2019plan\restock-report.xlsx',sheet_name="restock-report",startrow=0,header=True,index=False)

        data_shipped_CA.to_excel(r'D:\2019plan\Canada在途库存.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
        print("CA在途更新完成")
        
    elif key[2]in file:
        print("有MX在途")           
        data_tsv5['到货日期']=""
        data_tsv5['周数']=""
        print("在途库存列数比较",data_tsv5.columns,salescolumns_MX)
        data_tsv5.columns=salescolumns_MX
        
    
        print(data_tsv5)
        data_shipped_MX=data_shipped_MX.append(data_tsv5,ignore_index=True)
                     
       #追加到在途计划 data_csv2.to_excel(r'D:\2019plan\restock-report.xlsx',sheet_name="restock-report",startrow=0,header=True,index=False)

        data_shipped_MX.to_excel(r'D:\2019plan\Mexico在途库存.xlsx', sheet_name="Sheet1",startrow=0,header=True,index=False)
        print("MX在途更新完成")

    
    else:
        print("什么在途文件都没有")

########################################################################################################################################################################
##以下为运行新站的计划
##以下为运行新站的计划
##以下为运行新站的计划
print("以下为运行新站的计划")
print("以下为运行新站的计划")
        
# -*- coding:utf-8 –*-
import os
import pandas as pd
import shutil

# -*- coding:utf-8 –*-
import os
import pandas as pd
import shutil



#导入库存

src_dir_path_inventory=r'D:\运营\计划数据\Newcountries\当日库存'

key =['NEW-US','CA','MX','UK','IT','DE','JP','ES','FR','HM-US']
 
#获取原来库存文件的列名
data_inventory_US=pd.read_excel(r'D:\SailingstarFBA计划\当日Amazon库存.xlsx')
data_inventory_CA=pd.read_excel(r'D:\SailingstarFBA计划\Canada当日Amazon库存.xlsx')
data_inventory_UK=pd.read_excel(r'D:\SailingstarFBA计划\UK当日Amazon库存.xlsx')
data_inventory_IT=pd.read_excel(r'D:\SailingstarFBA计划\IT当日Amazon库存.xlsx')
data_inventory_DE=pd.read_excel(r'D:\SailingstarFBA计划\DE当日Amazon库存.xlsx')           
data_inventory_MX=pd.read_excel(r'D:\SailingstarFBA计划\MX当日Amazon库存.xlsx')
data_inventory_JP=pd.read_excel(r'D:\SailingstarFBA计划\JP当日Amazon库存.xlsx')
                                
inventorycolumns_US=data_inventory_US.columns.tolist()
inventorycolumns_CA=data_inventory_CA.columns.tolist()
inventorycolumns_MX=data_inventory_MX.columns.tolist()
inventorycolumns_UK=data_inventory_UK.columns.tolist()
inventorycolumns_DE=data_inventory_DE.columns.tolist()
inventorycolumns_IT=data_inventory_IT.columns.tolist()
inventorycolumns_JP=data_inventory_JP.columns.tolist()

print(inventorycolumns_US)

# 遍历文件夹

for file in os.listdir(src_dir_path_inventory):
    
    print(os.listdir(src_dir_path_inventory))
    
    data_csv = pd.read_csv(r'D:\\运营\\计划数据\\Newcountries\\当日库存\\'+ str(file),encoding='Latin1')
    # 读文件
#
    #US
    if key[0] in file:
        print(file)
    # 执行语句
        print("有US库存")
       
         # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        print(data_csv)
        print("比较列",data_csv.columns,data_inventory_US.columns)
        data_csv.columns=inventorycolumns_US                      
        data_csv.to_excel(r'D:\SailingstarFBA计划\当日Amazon库存.xlsx',sheet_name="US-new24374599305018570",startrow=0,header=True,index=False)
        
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
        
    elif key[1]in file:
        print("有CA库存")
 
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        print("比较列",data_csv.columns,data_inventory_CA.columns)
        data_csv.columns=inventorycolumns_CA    
                         
        data_csv.to_excel(r'D:\SailingstarFBA计划\Canada当日Amazon库存.xlsx', sheet_name="当前Amazon库存",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
        
      
    elif key[2]in file:
        print("有MX库存")
   
        #df_data.columns.tolist())
        print(data_csv)
        print("bijiaolie",data_csv.columns,data_inventory_MX.columns)
        data_csv.columns=inventorycolumns_MX                      
        data_csv.to_excel(r'D:\SailingstarFBA计划\MX当日库存.xlsx',sheet_name="24493532708018574",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
        
    elif key[3]in file:
        print("有UK库存")
  
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        print("bijiaolie",data_csv.columns,data_inventory_UK.columns)
        data_csv.columns=inventorycolumns_UK    
                         
        data_csv.to_excel(r'D:\SailingstarFBA计划\UK当日Amazon库存.xlsx', sheet_name="UK25372824608018570",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
    elif key[4]in file:
        print("有IT库存")
    
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        print("bijiaolie",data_csv.columns,data_inventory_IT.columns)
        data_csv.columns=inventorycolumns_IT  
                         
        data_csv.to_excel(r'D:\SailingstarFBA计划\IT当日Amazon库存.xlsx', sheet_name="当前Amazon库存",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
    elif key[5]in file:
        print("有DE库存")
        
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist()
        print("bijiaolie",data_csv.columns,data_inventory_DE.columns)
        data_csv.columns=inventorycolumns_DE   
                         
         # data_csv.to_excel(r'D:\SailingstarFBA计划\DE当日Amazon库存', sheet_name="当前Amazon库存",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
    elif key[6]in file:
        print("有JP库存")
       
        # 旧语句data_csv = pd.read_csv(r'D:\\运营\\计划数据\\老站\\当日库存\\'+str(file),encoding='utf-8 ', error_bad_lines=False)     # 读取以分
        
        #df_data.columns.tolist())
        print("bijiaolie",len(data_csv.columns),len(data_inventory_JP.columns))
        data_csv.columns=inventorycolumns_JP   
                         
        data_csv.to_excel(r'D:\SailingstarFBA计划\JP当日Amazon库存.xlsx', sheet_name="当前Amazon库存",startrow=0,header=True,index=False)
        shutil.move('D:\\运营\\计划数据\\Newcountries\\当日库存\\'+str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/当日库存')
        #将文件转移到历史文件夹

        
        
    else:
        print("没有US,CA,MX,UK,IT,DE,JP当日库存")
   
    
     
# 导入reStock
sheetnamedic={ "US" : 'US-restock-report' , 'CA' : 'CA-restock-report_12-22-2020_09' , 'MX' : 'MX-restock-report_','UK':'UK-restock-report','IT':'UK-restock-report','DE':'UK-restock-report','JP':'JP-restock-report_12-22-2020_09' }
src_dir_path_restock=r'D:\运营\计划数据\Newcountries\restock'
print(os.listdir(src_dir_path_restock))
for file in os.listdir(src_dir_path_restock):
    data_csv2 = pd.read_table(r'D:\\运营\\计划数据\Newcountries\\restock\\'+ str(file))    # 读取以分        
    for i in range(len(key)):
        if key[i] in file: 
            print(key[i])
            print(str(key[i]))
            data_csv2.to_excel(r'D:\SailingstarFBA计划\ '+str(key[i])+"-restock-report.xlsx",sheet_name=str(sheetnamedic[key[i]]),startrow=0,header=True,index=False)
            print("已导出"+str(key[i])+"-restock-report")
            break
        else:
            print("查看resock目标文件库，缺key列表国家的目标文件")
    shutil.move(r'D:\\运营\\计划数据\\Newcountries\\restock\\'+ str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/restock')   
         
        
#re导出完毕

# -*- coding:utf-8 –*-
import os
import pandas as pd
import shutil
import openpyxl
key =['US','CA','MX','UK','IT','DE','JP','ES','FR']
  
# 复制TSV在途库存
shippedfiledic={ 'US' : '在途库存' , 'CA' : 'Canada在途库存' , 'MX' : 'MX在途库存','UK':'UK在途库存','IT':'IT在途库存','DE':'DE在途库存','JP':'JP在途库存','ES':'ES在途库存' ,'FR':'FR在途库存'}
src_dir_path_shipped=r'D:\运营\计划数据\Newcountries\在途库存'
shippedfilepath=r'D:\\SailingstarFBA计划\\'
#数据源文件目录
print(os.listdir(src_dir_path_shipped))

a=0

for file in os.listdir(src_dir_path_shipped):
    #遍历数据源文件
                  
    #旧程序data_shipped_US=pd.read_excel(r'D:\SailingstarFBA计划\在途库存.xlsx')
    #data_shipped_CA=pd.read_excel(r'D:\SailingstarFBA计划\Canada在途库存.xlsx')
    #data_shipped_MX=pd.read_excel(r'D:\SailingstarFBA计划\Mexico在途库存.xlsx')
    #salescolumns_US=data_shipped_US.columns.tolist()
    #salescolumns_CA=data_shipped_CA.columns.tolist()
    #salescolumns_MX=data_shipped_MX.columns.tolist()
    data_tsv5 = pd.read_csv(r'D:/运营/计划数据/NewCountries/在途库存/'+ str(file), sep='\t',nrows =5)
    print(data_tsv5.iloc[0,1])
    batchnumber=data_tsv5.iloc[0,1]
    print(data_tsv5)
    print(batchnumber)
    data_tsv5= pd.read_csv(r'D:/运营/计划数据/NewCountries/在途库存/'+ str(file), sep='\t',header=6)
    print(data_tsv5)
 
    #batchnumber= data_tsv5.iat[0,1]
    #读取批次号
    if len(data_tsv5.columns)>10:
        x=[9,10]
        data_tsv5.drop(data_tsv5.columns[x], axis=1, inplace=True)
        print(data_tsv5)

    data_tsv5["批次"]=batchnumber
    data_tsv5['到货日期']=""
    data_tsv5['周数']=""
    print(data_tsv5)

    #读取源文件去掉前8行；可以用去掉前8行重写

    print(data_tsv5)
               
    #加入批次号作为一列；可以用assign重写
    b=0
    for i in range(len(key)):
    #遍历国家名字典
        if key[i] in file:
        #如果数据源文件名中包含国家
                    
            datashipped=pd.read_excel(shippedfilepath+ shippedfiledic[key[i]]+'.xlsx' )
            print(datashipped)
            #读取国家i的目标文件
            print(key[i])
                       
            data_tsv5.columns=datashipped.columns                            
            datashipped=datashipped.append(data_tsv5,ignore_index=True)
            datashipped.to_excel(r'D:\\SailingstarFBA计划\\'+shippedfiledic[key[i]]+'.xlsx',sheet_name="Sheet1",startrow=0,header=True,index=False)
            print("已导出"+str(key[i])+"在途库存导入")
            a+=1
            
            #将源数据文件加到目标文件                          
            #else
            #print("查看resock目标文件库，缺key列表国家的目标文件"）    
            break
            #一旦符合条件后面就不循环找了;实际就是找到目标文件中的第一个国家就跳出。
        else:
            b+=1
            
        if b==len(key):
            print("检查源文件")     
        
      
    

         
    
    shutil.move(r'D:\\运营\\计划数据\\Newcountries\\在途库存\\'+ str(file), 'D:/运营/HistoricalData/计划数据/Newcountries/在途库存')
   
    
print("完成复制在途库存，完成了"+str(a)+"个在途库存导入")
      
#完成复制在途库存；可以写成一个def函数

#复制销售数据

samplecolumn=pd.read_excel('D:\\运营\\计划数据\\NewCountries\\samplecolumnUSMX.xlsx')
#取得墨西哥和美国的标准列名
USMXColumnlist=samplecolumn.columns.tolist()
key =['US','CA','MX','UK','IT','DE','JP','ES','FR']
#开始导入销售数据
#思路：1.遍历销售数据2. 导入file1加上国家，if 如果是美国，墨西哥（或者如果列数为固定的）则插入2个字段 ;将file1 的dataframe追加到 目标文件dataframe ;导出目标dataframe到excel
#思路：1.遍历销售数据2. 导入file1加上国家，if 如果是美国，墨西哥（或者如果列数为固定的）则插入2个字段 ;将file1 的dataframe追加到 目标文件中间件dataframe ;导出目中间件dataframe到目标dataframe到excel

salesfilepath = 'D:\\运营\\计划数据\\NewCountries\\销售数据\\'
Newallsales=pd.read_excel('D:/SailingstarFBA计划/NEW-ALL周销售数据.xlsx')
NewColumnlist= Newallsales.columns.tolist()
#取得目标文件的列名
               
a=0
for salesfile in os.listdir(salesfilepath):
    print(salesfile)
#遍历数据文件
    sourcedata=pd.read_csv(salesfilepath+str(salesfile)).assign(Country=os.path.basename(salesfile).split('_')[0], 日期=os.path.basename(salesfile).split('_')[1])
    sourcedata['日期'] = pd.to_datetime(sourcedata['日期'])
    print(sourcedata[['Country','日期']])
    for i in range(len(key)):
    #遍历国家名     
        if key[i] in salesfile:
                       
            if len(sourcedata.columns)==15:
            #如果列数=14
                #USMXColumnlist =["(Parent) ASIN","(Child) ASIN","Title","SKU","Sessions","Session Percentage","Page Views","Page Views Percentage","Buy Box Percentage","Units Ordered","Unit Session Percentage","Ordered Product Sales","Total Order Items","Country","日期"]
                #定义标准列名，给美国和MX赋值列名，reindex列名

                print(USMXColumnlist)
                 

                sourcedata.columns=USMXColumnlist
             
                #给数据源文件列名赋标准值
                
                print(NewColumnlist)
                sourcedata=sourcedata.reindex(columns=NewColumnlist)
                #给源文件插入全部缺失的列名，行值为0
                print(len(sourcedata.columns))
                #test.to_excel('D:/运营/计划数据/NewCountries/销售数据/temp.xlsx')
            elif len(sourcedata.columns)==19:
                sourcedata['周数']=""
                print(len(sourcedata.columns),len(NewColumnlist))
                sourcedata.columns=NewColumnlist
                 ##给数据源文件列名赋目标文件列名标准值
            else:
                  print("检查销售数据文件的列名")


            Newallsales=Newallsales.append(sourcedata,ignore_index=True)
            print(Newallsales)
             #追加数据到目标文件
            print("添加销售数据"+str(key[i]))
            break
    shutil.move(r'D:\\运营\\计划数据\\Newcountries\\销售数据\\'+ str(salesfile), 'D:/运营/HistoricalData/计划数据/Newcountries/销售数据')
    a+=1

#Newallsales["日期"] = Newallsales["日期"].dt.strftime("%Y-%m-%d")
#统一Newallsales的日期格式但是报错


maxtime2=pd.to_datetime(Newallsales['日期'].max())
  
print("最晚时间",maxtime2)
Newallsales['日期']=pd.to_datetime(Newallsales['日期'])
#获取最晚时间
Newallsales['周数']=(maxtime2-Newallsales['日期']).dt.days//7+1
#给周数赋值
Newallsales.to_excel(r'D:\SailingstarFBA计划\NEW-ALL周销售数据.xlsx',sheet_name="Sheet1",startrow=0,header=True,index=False)

print("导出目标文件到Excel，数量为"+str(a))
    
#销售数据复制完毕



