# 使用Oenyxl编写，没有使用pandas
### 第一部分汇总各国bulk广告数据到汇总表
#用于实际使用，从零开始导入10周最新的。创建新的文件，存到指定文件名。：测试OK。

# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表


print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )

newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)


#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#？？？旧的程序，不要了bulkoperationworkbook=load_workbook(r'D:\运营\bulkoperation模板.xlsx')
#???sheetbulkoperation=bulkoperationworkbook.worksheets[0]

#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'

for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name=1).assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
    sourcedata.replace(",",".",inplace=True)
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=1
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')#汇总所有广告数据

 
    #???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')

#拷贝一份sourcedata到bulkoperation文件夹,文件名不变
    shutil.copy(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\\运营\\bulkoperationfiles\\')
#移动广告数据到历史数据 
    shutil.move(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile),r'D:\\运营\\HistoricalData\\周bulk广告数据\\')
    
#取得日期列的值的列表：取得最大值：  取得日期列的第一个值，计算周数：在周数的指定位置写入周数。


for i in range(2,sheet.max_row+1):
    b=sheet.cell(row=i,column=30).value

    bnewtime=sheet.cell(row=i,column=30).value

    c=(maxtime-bnewtime).days//7+1

    sheet.cell(row=i, column=31).value =c
    
                      
    
#保存所有广告数据汇总表
bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表.xlsx')
###???bulkoperationworkbook.save(r'D:\运营\bulkoperation.xlsx')


#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

#####以下为处理bulk操作报表的程序#####以下为处理bulk操作报表的程序

 
print("以下为处理bulk操作报表的程序")

##先用临时文件测试后copy来
#此程序测试完后应该考入bulkoperation程序中

print("以下为处理bulk操作报表的程序")



#定义bulk数据汇总表所在路径
Allbulkpath='D:\\运营\\'
Allbulk=pd.read_excel(Allbulkpath+'周bulk广告数据汇总表.xlsx')


 

#####生成各种Campaign summary

 

AllbulkCampaign=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
#AllbulkCampaign1week=Allbulk[(Allbulk['Record Type']=="Campaign")&(Allbulk['周数']==1)].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')


AllbulkCampaignWEEK=Allbulk[Allbulk['Record Type']=="Campaign"].groupby(["Country","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignSKUWEEK=Allbulk[Allbulk['Record Type']=="Ad"].groupby(["Country","SKU","Campaign","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')
AllbulkCampaignKeywordWEEK=Allbulk.groupby(["Country","Campaign","Keyword or Product Targeting","Ad Group","周数"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")

writer=pd.ExcelWriter(Allbulkpath+'周bulk数据Summary.xlsx')
AllbulkCampaign.to_excel(writer,"Campaign汇总")
AllbulkCampaignWEEK.to_excel(writer,"CampaignWEEK汇总")
AllbulkSKUWEEK.to_excel(writer,"SKU-WEEK")
AllbulkCampaignSKUWEEK.to_excel(writer,"SKU-Campaign-WEEK")
AllbulkCampaignKeywordWEEK.to_excel(writer,"Keyword-Campaign-WEEK")
writer.save()

################生成修改bulk所使用的汇总Summary表

input("zanting")
#！！！！筛选汇总表的数据---后续可以按日期>某个日期来筛选
AllbulkD5=Allbulk[Allbulk['Keyword or Product Targeting'].notna()]

# AllbulkCampaignKeyword1Week=Allbulk[(Allbulk['Keyword or Product Targeting'].notna())&(Allbulk['周数']==1)].groupby(["Country","Campaign"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg('sum')

#AbulkCampaignKeyword1week["zhuanhualv"]=AllbulkCampaignKeyword1week['Orders']/AllbulkCampaignKeyword1week['Clicks']


AllbulkCampaignKeyword=AllbulkD5.groupby(["Country","Campaign","Keyword or Product Targeting"],as_index=False)[['Impressions','Clicks','Spend','Orders','Total Units','Sales']].agg("sum")


#print(AllbulkCampaignKeyword)
maxrow=len(AllbulkCampaignKeyword)
#maxrow1Week=len(AllbulkCampaignKeyword1Week)
AllbulkCampaignKeyword["zhuanhualv"]=AllbulkCampaignKeyword['Orders']/AllbulkCampaignKeyword['Clicks']
print(maxrow)
print(AllbulkCampaignKeyword)

AllbulkCampaignKeyword.to_excel(Allbulkpath+'周bulk广告数据汇总表huizong.xlsx')


#################                     
#bulkoperation=pd.read_excel(Allbulkpath+'bulkoperation.xlsx')#注意不是bulk模板
#bulkoperation["是否更新"]=""
#bulkoperationmaxrow=len(bulkoperation)



#从国家-bulkopration遍历
bulkoperationfilespath = 'D:\\运营\\bulkoperationfiles\\'
bulkoperationfilesnewpath='D:\\运营\\bulkoperationfilesNEW\\'
bulkfilecountrylist=[]
for bulkoperationfile in os.listdir(bulkoperationfilespath):
    clickatempt=10
    clickenough=16
    #提取国家名
    bulkfileCountry=os.path.basename(bulkoperationfile).split('_')[0]
  
    print(bulkfileCountry,type(bulkfileCountry))
  
    
    bulkfilecountrylist.append(bulkfileCountry)
   
    print(bulkfilecountrylist) #??后续可以把处理的国家写到一个处理报告里

    dfbulkfile=pd.read_excel(bulkoperationfilespath+str(bulkoperationfile),engine="openpyxl",sheet_name=1)
    dfbulkfile.replace(",",".",inplace=True)
    print(dfbulkfile)                            
    ##？？是否可以使用bulkoperationfile这一相同变量名？

##？？如何考察Camaign 1. 整体 2. 里面的关键词是什么状态：所以还是先从Allbulcampaignword遍历
############################ 从Allbulcampaignword遍历

             #for i in range(0,maxrow):
    #定义主要指标                               
    #clicks=AllbulkCampaignKeyword.iloc[[i],[4]].values[0][0]
    #zhuanhualv=AllbulkCampaignKeyword.iloc[[i],[9]].values[0][0]
    #keyword=AllbulkCampaignKeyword.iloc[[i],[2]].values[0][0]                                    
#   打开变更记录表
    bulkrecordpath = 'D:\\运营\\bulk变更记录\\'
    if not os.listdir(bulkrecordpath):
        print("文件夹为空")
    
        dfbulkmodifyrecord=pd.DataFrame()
        print(dfbulkmodifyrecord)
        dfbulkmodifyrecord=dfbulkmodifyrecord.append(dfbulkfile)
        print("新的记录表",dfbulkmodifyrecord)
    else: 
 
        print("文件夹是有文件",os.listdir(bulkrecordpath))
        #遍历记录文件夹
        for bulkmodifyrecordfile in os.listdir(bulkrecordpath):#
            print("os:",os.path.basename(bulkmodifyrecordfile).split('_')[0],"str:",str(os.path.basename(bulkmodifyrecordfile).split('_')[0]))
            
            if os.path.basename(bulkmodifyrecordfile).split('_')[0]==str(bulkfileCountry):
               yanzhnegfu=1
               bulkmodifyrecordfilefound=os.path.basename(bulkmodifyrecordfile)
                
               print("有匹配os.path.basename(bulkmodifyrecordfile)")
               break
        if yanzhnegfu==1:
            print("文件夹内有变更记录")
            dfbulkmodifyrecord=pd.read_excel(bulkrecordpath+str(bulkmodifyrecordfilefound),engine="openpyxl",sheet_name=0)
            print(dfbulkmodifyrecord)
            dfbulkmodifyrecord=pd.concat([dfbulkmodifyrecord,dfbulkfile])
            dfbulkmodifyrecord.drop_duplicates(subset=["Record Type","Campaign","Campaign Targeting Type","Product Targeting ID","Ad Group","Keyword or Product Targeting","Product Targeting ID" ,"Match Type",  "SKU","Bidding strategy","Placement Type" ],inplace=True)
        else:
             
            print("文件夹内没有变更记录")
        
            dfbulkmodifyrecord=pd.DataFrame()
            print(dfbulkmodifyrecord)
            dfbulkmodifyrecord=dfbulkmodifyrecord.append(dfbulkfile)
            print("新的记录表",dfbulkmodifyrecord)
        
  
      
         
    Today= datetime.datetime.today().strftime('%Y-%m-%d')
    Columnnew1=Today+"_"+"最新状态"
    Columnnew2=Today+"_"+"更新内容"
    print(Today,Columnnew1,Columnnew2)
    dfbulkmodifyrecord[Columnnew1]=""
    dfbulkmodifyrecord[Columnnew2]=""
    print(dfbulkmodifyrecord.columns)
    
#########################
  
######定义各种判断条件：用筛选的方式
    #####可以自合df筛选条件的字符串                                
    bulcountryconditionstr='AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"=="bulkfileCountry")'                            
   
    bulkcountrycondionclk1str='(AllbulkCampaignKeyword["Clicks"]>(clickatempt-1)) &(AllbulkCampaignKeyword["Clicks"]<(clickenough+1))]'
    bulkcountrycondionclk21str='AllbulkCampaignKeyword["Clicks"]>clickenough'
    bulkcountrycondionzhl01str= '(AllbulkCampaignKeyword["zhuanhualv"]==0]'
    bulkcountrycondionzhl02str= '(AllbulkCampaignKeyword["zhuanhualv"]>0.2]'
    bulkcountrycondionzhl032str='(AllbulkCampaignKeyword["zhuanhualv"]>0.5]'                                 
    #bulkconditiondic={}#
##1 在AllbulkCampaignKeyword筛选出第一种操作的对象条件表:点击大于9小于17,转化率=0
    #做了reindex
    bocondition01df=AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"]==str(bulkfileCountry)) &(AllbulkCampaignKeyword["Clicks"]>(clickatempt-1)) &(AllbulkCampaignKeyword["Clicks"]<(clickenough+1))&(AllbulkCampaignKeyword["zhuanhualv"]==0)].reset_index(drop=True)
    
   
    print("bocondition01df",bocondition01df)
  
    
###遍历这个符合条件1的条件表
                                                                                                    
    for boi in range(0,len(bocondition01df)):
                                                                                                    
        boi_keyword=bocondition01df.iloc[[boi],[2]].values[0][0]
        print(boi_keyword,boi)
         
        boi_campaign=bocondition01df.iloc[[boi],[1]].values[0][0]                                                                                            
                                                                                                    

#将bulkfile中这个campaign-词的maxbid批量进行修改 
                                                                                           
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Max Bid"]=dfbulkfile["Max Bid"]*0.5
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"更新内容"]="点击大于9小于17,转化率=0，降价0.5"             
        print("01:将bulkfile中这个campaign-词的maxbid进行修改")
        
#将bocondition01df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew1]="bocondition01df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew2]="点击大于9小于17,转化率=0，降价0.5" 

 
###2 筛选出第2种操作的对象条件表：点击》16，转化率为0
    bocondition02df=AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"]==bulkfileCountry) &(AllbulkCampaignKeyword["Clicks"]>clickenough)&(AllbulkCampaignKeyword["zhuanhualv"]==0)]
    print(bocondition02df)
    for boi in range(0,len(bocondition02df)):###遍历这个符合条件2的条件表
                                                                                                    
        boi_keyword=bocondition02df.iloc[[boi],[2]].values[0][0]
        boi_campaign=bocondition02df.iloc[[boi],[1]].values[0][0]
#将bulkfile中这个campaign-词的Status批量进行修改
                                            
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Status"]="paused"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"更新内容"]="点击>16，转化率为0,paused"

        print("02:将bulkfile中这个campaign-词的Status进行修改")
#将bocondition02df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign)&(dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew1]="bocondition02df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign)&(dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew2]="点击>16，转化率为0,paused"
        print("02:将bulkfile中这个campaign-词的修改记录" )



                        
###3 筛选出第3种操作的对象条件表：3. 累计点击>16次且 0.0625<转化率<=0.125 (16次小于2单）： 点击费用*0.75
    bocondition03df=AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"]==bulkfileCountry) &(AllbulkCampaignKeyword["Clicks"]>clickenough)&(AllbulkCampaignKeyword["zhuanhualv"]>0.0625)&(AllbulkCampaignKeyword["zhuanhualv"]<0.125)]               
    for boi in range(0,len(bocondition03df)):###遍历这个符合条件3的条件表
                                                                                                    
        boi_keyword=bocondition03df.iloc[[boi],[2]].values[0][0]
        boi_campaign=bocondition03df.iloc[[boi],[1]].values[0][0]

        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Max Bid"]=dfbulkfile["Max Bid"]*0.75
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"更新内容"]="点击>16，转化率低于0.125, 降低价*0.75"
        print("03: 累计点击>16次且 0.0625<转化率<=0.125 (16次小于2单）： 点击费用*0.75")
#将bocondition03df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew1]="bocondition03df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew2]="点击>16，转化率低于0.125, 降低价*0.75"


###4 筛选出第4种操作的对象条件表：4. 累计点击>16次且  转化率<0.0626 (16次小于2单）： 点击费用*0.33
    bocondition04df=AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"]==bulkfileCountry) &(AllbulkCampaignKeyword["Clicks"]>clickenough)&(AllbulkCampaignKeyword["zhuanhualv"]<0.0625)&(AllbulkCampaignKeyword["zhuanhualv"]>0)]
    for boi in range(0,len(bocondition04df)):
           #定义要找的词和Campaign                                   
        boi_keyword=bocondition04df.iloc[[boi],[2]].values[0][0]
        boi_campaign=bocondition04df.iloc[[boi],[1]].values[0][0]
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),'Max Bid']=dfbulkfile["Max Bid"]*0.5
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"更新内容"]="点击>16，转化率低于0.0625, 降低价*0.5"                                            

        print("04:累计点击>16次且  转化率<0.0626 (16次小于2单）： 点击费用*0.5")
#将bocondition04df修改记录写入记录表        
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew1]="bocondition04df"

        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew2]="点击>16，转化率低于0.0625, 降低价*0.5" 
###5 筛选出第4种操作的对象条件表转化率》0.2的一律开启
    bocondition05df=AllbulkCampaignKeyword[(AllbulkCampaignKeyword["Country"]==bulkfileCountry) &(AllbulkCampaignKeyword["zhuanhualv"]>0.2)]
    for boi in range(0,len(bocondition05df)):
           #定义要找的词和Campaign                                   
        boi_keyword=bocondition05df.iloc[[boi],[2]].values[0][0]
        boi_campaign=bocondition05df.iloc[[boi],[1]].values[0][0]
           # 打开Campaign enabled                                    
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Campaign"),"Campaign Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Campaign"),"更新内容"]="转化率>0.2，开启"
#将bocondition05df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Campaign"),Columnnew1]="bocondition05df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Campaign"),Columnnew2]="转化率>0.2，开启"
        
           # 打开Ad Group enabled
                                   
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad Group"),"Campaign Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad Group"),"更新内容"]="转化率>0.2，开启"                                   
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad Group"),"Ad Group Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad Group"),"更新内容"]="转化率>0.2，开启"

#将bocondition05df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Ad Group"),Columnnew1]="bocondition05df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Ad Group"),Columnnew2]="转化率>0.2，开启"


        
           # 打开Ad enabled                                    
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"Campaign Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"更新内容"]="转化率>0.2，开启"                                    
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"Ad Group Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"更新内容"]="转化率>0.2，开启"                                  
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Record Type']=="Ad"),"更新内容"]="转化率>0.2，开启"
#将bocondition05df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Ad"),Columnnew1]="bocondition05df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Record Type']=="Ad"),Columnnew2]="转化率>0.2，开启"
           #打开keyword enabled                                     
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Ad Group Status"]="enabled"
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"Campaign Status"]="enabled"
                                             
        dfbulkfile.loc[(dfbulkfile["Campaign"]==boi_campaign) & (dfbulkfile['Keyword or Product Targeting']==boi_keyword),"更新内容"]="转化率>0.2，开启"
#将bocondition05df修改记录写入记录表
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew1]="bocondition05df"
        dfbulkmodifyrecord.loc[(dfbulkmodifyrecord["Campaign"]==boi_campaign) & (dfbulkmodifyrecord['Keyword or Product Targeting']==boi_keyword),Columnnew2]="转化率>0.2，开启"
        
           ##?? 后续价格按推荐值调整                                   
        print("05:转化率》0.2的一律开启")
#########a将修改过的文件存入bulkoperationfilesnewpath
    dfbulkfile.to_excel(bulkoperationfilesnewpath+"Re_"+Today+"_"+str(bulkoperationfile),index=False)
#########a将更改记录文件存入bulkrecordpath    
    dfbulkmodifyrecord.to_excel(bulkrecordpath+bulkfileCountry+"_"+"更改记录"+".xlsx",index=False)
########将周bulk广告数据里的文件移动到历史文件夹    
    #shutil.move(r'D:\\运营\\周bulk广告数据\\'+ str(bulkdatafile), r'D:\运营\HistoricalData\周bulk广告数据')

#########3将bulkoperation里的文件删除
    os.remove(bulkoperationfilespath+ str(bulkoperationfile))
########或者可以移动将bulkoperation里的文件移动到历史文件夹
    #shutil.move(bulkoperationfilespath+ str(bulkdatafile), r'D:\运营\HistoricalData\周bulk广告数据\bulkoperationfiles')
 
