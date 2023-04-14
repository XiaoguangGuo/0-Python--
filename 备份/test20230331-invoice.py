import pandas as pd
from openpyxl import load_workbook

# 读取Excel文件
filename = r'D:\\运营\\1数据源\\invoice\\shipment.xlsx'
wb = load_workbook(filename)
sheet = wb.active

# 初始化变量和结果dataframe
shipment_data = {
    "Shipment ID": [],
    "Box ID": [],
    "SKU": [],
    "数量": [],
    "Weight": [],
    "Unit": [],
    "Length": [],
    "Width": [],
    "Height": [],
    "UnitDimensions": []
}
shipment_id = ""
box_id = ""

# 遍历第一列
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    if cell_value is not None:
        cell_value = str(cell_value)  # 将cell_value转换为字符串

        if "Shipment ID:" in cell_value:
            shipment_id = cell_value.replace("Shipment ID:", "").strip()

        elif "Box ID:" in cell_value:
            box_id = cell_value.replace("Box ID:", "").strip()

            # 寻找SKU和数量
            sku_row = row + 1
            while sku_row <= sheet.max_row:
                sku_value = str(sheet.cell(row=sku_row, column=1).value)
                if "SKU:" in sku_value:
                    sku = sku_value.replace("SKU:", "").strip()
                    quantity_row = sku_row + 1
                    quantity = None
                    while quantity_row <= sheet.max_row:
                        try:
                            quantity = int(sheet.cell(row=quantity_row, column=1).value)
                            break
                        except (ValueError, TypeError):
                            quantity_row += 1

                        # 如果当前行是最后一行，寻找下一个有效的数字作为数量
                        if quantity_row == sheet.max_row and quantity is None:
                            search_row = sheet.max_row
                            while search_row > 0:
                                try:
                                    quantity = int(sheet.cell(row=search_row, column=1).value)
                                    break
                                except (ValueError, TypeError):
                                    search_row -= 1

                    weight_row = row + 1
                    dimensions_row = row + 1
                    while weight_row < sheet.max_row:
                        weight_value = str(sheet.cell(row=weight_row, column=1).value)
                        if "weight:" in weight_value.lower():
                            weight = sheet.cell(row=weight_row + 1, column=1).value
                            unit = sheet.cell(row=weight_row + 2, column=1).value
                            break
                        weight_row += 1

                    while dimensions_row < sheet.max_row:
                        dimensions_value = str(sheet.cell(row=dimensions_row, column=1).value)
                        if "dimensions:" in dimensions_value.lower():
                            length = sheet.cell(row=dimensions_row + 1, column=1).value
                            width = sheet.cell(row=dimensions_row + 3, column=1).value
                            height = sheet.cell(row=dimensions_row + 5, column=1).value
                            unit_dimensions = sheet.cell(row=dimensions_row + 6, column=1).value
                            break
                        dimensions_row += 1

                    shipment_data["Shipment ID"].append(shipment_id)
                    shipment_data["Box ID"].append(box_id)
                    shipment_data["SKU"].append(sku)
                    shipment_data["数量"].append(quantity)
                    shipment_data["Weight"].append(weight)
                    shipment_data["Unit"].append(unit)
                    shipment_data["Length"].append(length)
                    shipment_data["Width"].append(width)
                    shipment_data["Height"].append(height)
                    shipment_data["UnitDimensions"].append(unit_dimensions)

                    sku_row = quantity_row  # 更新sku_row以跳过已处理的数量行
                elif "Box ID:" in sku_value:
                    break  # 跳出循环以处理下一个Box ID
                
                else:
                    sku_row += 1

shipment_df = pd.DataFrame(shipment_data)
print(shipment_df)
shipment_df.to_excel(r'D:\\运营\\invoicedraft.xlsx')



# 读取产品信息表格
# 读取产品信息表格
products_df = pd.read_excel(r'G:\\运营\invoice\\发票基础信息表.xlsx')
products_df['SKU'] = products_df['SKU'].fillna('')

# 读取另一个 DataFrame
other_df = shipment_df

# 遍历 other_df 的 SKU 列，在 products_df 的产品 SKU 列中找到包含对应值的单元格，将这一行对应的信息合并到 other_df 中
for index, row in other_df.iterrows():
    for sku in row['SKU'].split():
        matching_rows = products_df[products_df['SKU'].str.contains(sku)]
        if not matching_rows.empty:
            matching_row = matching_rows.iloc[0]
            for col_name in ['产品SKU', '英文产品名称', '中文产品名称', '单个产品申报价值USD', "HSCODE",
                             'Brand(品牌)*', 'Model（型号）*', 'Material（材质）*',
                             'Purpose(用途)*', '是否带电（Y/N）', 'PICTURES（图片）*',
                             '产品销售链接', '内部名称', '产品销售价格', '英文材质',
                             '英文用途']:
                other_df.at[index, col_name] = matching_row[col_name]
            other_df.at[index, 'SKU'] = matching_row['SKU']
            



# 修改箱尺寸列
other_df["箱尺寸"] = other_df["Length"].astype(str) + "*" + other_df["Width"].astype(str) + "*" + other_df["Height"].astype(str)

# 计算箱体积并保留小数后三位
other_df["箱体积"] = (other_df["Length"] * other_df["Width"] * other_df["Height"] / 1000000).round(3)

# 增加箱数列
other_df["箱数"] = 1

# 修改中英文材质列
other_df["中英文材质"] = other_df["Material（材质）*"] + ' ' + other_df["英文材质"]

# 合并中英文用途列
other_df["中英文用途"] = other_df["Purpose(用途)*"] + other_df["英文用途"]

# 计算总申报价
other_df["总申报价"] = other_df["单个产品申报价值USD"] * other_df["数量"]# 将合并后的 DataFrame 写入 Excel 文件
other_df.to_excel(r'G:\\运营\invoice\\invoicedraft2.xlsx')

# 打印合并后的 DataFrame
print(other_df)
