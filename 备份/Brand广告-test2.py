
# -*- coding:utf-8 –*-
from openpyxl import load_workbook,Workbook
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime 
import shutil


#####1.将各国的bulk报表累积汇总到周bulk广告数据汇总表

print("请确认老站计划数据已经就位")
print("请确认新站计划数据已经就位")
print("使用方法：请确认已将最新的Bulk广告数据放到数据文件夹中,程序第一次运行请输入当天日期？")
a=input('按任意键继续，如未准备好则取消程序执行', )


newdate=input('输入最新日期y-m-d',) #输入最新一周的日期
maxtime=datetime.datetime.strptime(newdate,'%Y-%m-%d')
print(newdate)
print(maxtime)


#导入历史汇总表（可以汇总历史所有情况）
bulkhzWorkbook=load_workbook(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')


sheetnames=bulkhzWorkbook.sheetnames
print("品牌表名",sheetnames)
print(bulkhzWorkbook.sheetnames)
sheet=bulkhzWorkbook[sheetnames[0]]
print(sheet.title)
print(sheet.max_row,sheet.max_column)
#指定来源文件
bulkdatafilepath = 'D:\\运营\\周bulk广告数据\\'

for bulkdatafile in os.listdir(bulkdatafilepath):
    print(bulkdatafile)  
    datadate=bulkdatafile.split('-')[4]
    print(datadate)
    datatimedatetime=datetime.datetime.strptime(datadate,'%Y%m%d')
    print(datatimedatetime)                                            
    delta=(maxtime-datatimedatetime).days//7+1
    print(delta)
    
    sourcedata=pd.read_excel(bulkdatafilepath+str(bulkdatafile),engine="openpyxl",sheet_name="Sponsored Brands Campaigns").assign(Country=os.path.basename(bulkdatafile).split('_')[0], 日期=os.path.basename(bulkdatafile).split('-')[4])
    sourcedata=sourcedata.dropna(axis=0, how='all', thresh=None, subset=None, inplace=False)#只有全部为空才回被删除
    
    sourcedata['日期']=pd.to_datetime(sourcedata['日期'])
    sourcedata['周数']=delta
    for row in dataframe_to_rows(sourcedata,index= False,header= False): #使用这种方法很简单，但是日期是4位数字的文本，后续计算的时候要变更格式。
       
        sheet.append(row) #将来源文件写入目标文件
        ##???sheetbulkoperation.append(row)
    bulkhzWorkbook.save(r'D:\运营\周bulk广告数据汇总表-品牌.xlsx')#汇总所有广告数据

 
